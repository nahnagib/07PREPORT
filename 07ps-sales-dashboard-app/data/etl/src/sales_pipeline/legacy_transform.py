from __future__ import annotations

from dataclasses import dataclass, replace
import datetime as dt
import hashlib
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Union
import re

import numpy as np
import pandas as pd
from pandas import ExcelFile

PathLike = Union[str, Path]
DataFrame = pd.DataFrame
Series = pd.Series


# ============================================================
# Configuration
# ============================================================
@dataclass(frozen=True)
class PipelinePaths:
    base_dir: Path
    sales_path: Path
    targets_path: Path
    master_path: Path
    offdays_path: Path
    products_path: Path
    blocked_customers_path: Path
    output_path: Path

    @classmethod
    def from_base_dir(cls, base_dir: PathLike) -> "PipelinePaths":
        base = Path(base_dir)
        return cls(
            base_dir=base,
            sales_path=base / "script" / "odoo_quotations_20251102_1507.xlsx",
            targets_path=base / "Targets" / "sales_targets.xlsx",
            master_path=base / "SalesTeam.xlsx",
            offdays_path=base / "OffDays.xlsx",
            products_path=base / "PRODUCTS.xlsx",
            blocked_customers_path=base / "BlockedCustomers.xlsx",
            output_path=base / "Exports" / "SalesModel_OneOutput.xlsx",
        )


@dataclass(frozen=True)
class PipelineSettings:
    offdays_country: Optional[str] = "Libya"
    weekly_rest_day_name: str = "Friday"
    unknown_team_key: str = "MJ-TIK-00"
    unknown_salesperson_key: int = 0
    unknown_channel_name: str = "Unknown"
    unknown_channel_key: int = 1
    unknown_segment_name: str = "Unknown"
    export_product_conflicts: bool = True
    verbose: bool = True


@dataclass(frozen=True)
class PipelineConfig:
    paths: PipelinePaths
    settings: PipelineSettings


# ============================================================
# Logging
# ============================================================
class Logger:
    def __init__(self, enabled: bool = True) -> None:
        self.enabled = enabled

    def info(self, message: str) -> None:
        if self.enabled:
            print(f"[INFO] {message}")

    def ok(self, message: str) -> None:
        if self.enabled:
            print(f"[OK] {message}")

    def warn(self, message: str) -> None:
        if self.enabled:
            print(f"[WARN] {message}")


# ============================================================
# Constants
# ============================================================
SALES_COLUMN_MAP: Dict[str, str] = {
    "Order Date": "order_date",
    "Related Order": "order_number",
    "Customer": "customer",
    "Customer ID": "customer_id",
    "Customer Code": "customer_id",
    "Product": "product_name",
    "Salesperson": "salesperson",
    "Sales Team": "sales_team",
    "Company": "company",
    "Untaxed Total": "untaxed_total",
    "Total": "line_total",
    "Qty Invoiced": "quantity",
    "Status": "order_state",
    "Invoice Status": "invoice_status",
}

PEOPLE_SHEET_CANDIDATES = [
    "Salespersons", "Salesperson", "Sales People", "People", "salespersons", "salesperson"
]
TEAMS_SHEET_CANDIDATES = [
    "SalesTeams", "SalesTeam", "Teams", "Sales Team", "salesteam", "salesteams"
]


# ============================================================
# Helpers
# ============================================================
class TextUtils:
    UNKNOWN_PRODUCT_NAME = "Unknown Product"
    PRODUCT_CODE_PATTERN = re.compile(r"^(?:C\d+TE|S\d+|PRO|MT|\d+x\d+)$", flags=re.IGNORECASE)

    @staticmethod
    def norm(value: Any) -> Any:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return pd.NA
        text = str(value).strip()
        text = re.sub(r"\s+", " ", text)
        return pd.NA if text == "" else text

    @staticmethod
    def normalize_numeric_string(value: Any) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        if text == "":
            return ""
        text = re.sub(r"[^0-9,\.\-]", "", text)
        if "," in text and "." not in text:
            text = text.replace(",", ".")
        elif "," in text and "." in text:
            text = text.replace(",", "")
        return text

    @staticmethod
    def normalize_company(value: Any) -> str:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return ""
        text = str(value).strip()
        if text == "":
            return ""
        upper = text.upper()
        if "MAJA" in upper:
            return "Majaal"
        if "TIKA" in upper:
            return "Tika"
        return text

    @staticmethod
    def normalize_product_name(value: Any) -> str:
        return normalize_product_name(value)

    @staticmethod
    def _strip_invisible_product_chars(value: Any) -> str:
        if value is None or pd.isna(value):
            return ""
        text = str(value)
        text = text.replace("\u00a0", " ").replace("\u200b", "").replace("\u200c", "").replace("\u200d", "")
        text = re.sub(r"[\r\n\t]+", " ", text)
        return text

    @staticmethod
    def _has_arabic(text: str) -> bool:
        return bool(re.search(r"[\u0600-\u06FF]", text))

    @staticmethod
    def _has_latin_or_digit(text: str) -> bool:
        return bool(re.search(r"[A-Za-z0-9]", text))

    @classmethod
    def _display_token(cls, token: str) -> str:
        if cls._has_arabic(token) or not re.search(r"[A-Za-z]", token):
            return token
        wrapped = re.fullmatch(r"([^A-Za-z0-9]*)([A-Za-z0-9xX]+)([^A-Za-z0-9]*)", token)
        if wrapped:
            prefix, inner, suffix = wrapped.groups()
            if cls.PRODUCT_CODE_PATTERN.fullmatch(inner):
                inner_display = inner.lower() if re.fullmatch(r"\d+x\d+", inner, flags=re.IGNORECASE) else inner.upper()
                return f"{prefix}{inner_display}{suffix}"
            if inner.isupper() and len(inner) <= 4:
                return f"{prefix}{inner}{suffix}"
        if cls.PRODUCT_CODE_PATTERN.fullmatch(token):
            if re.fullmatch(r"\d+x\d+", token, flags=re.IGNORECASE):
                return token.lower()
            return token.upper()
        if token.isupper() and len(token) <= 4:
            return token
        family_case = {
            "xtreme": "Xtreme",
            "supercol": "SuperCol",
            "xtracol": "XtraCol",
            "germa": "Germa",
            "basalt": "Basalt",
            "travertine": "Travertine",
            "porcelain": "Porcelain",
            "marble": "Marble",
            "quartz": "Quartz",
        }
        lower = token.lower()
        return family_case.get(lower, lower[:1].upper() + lower[1:])

    @classmethod
    def normalize_product_display(cls, value: Any) -> str:
        text = cls._strip_invisible_product_chars(value)
        text = re.sub(r"\s+", " ", text).strip()
        if not text or text.lower() in {"nan", "none", "null", "<na>"}:
            return cls.UNKNOWN_PRODUCT_NAME

        text = re.sub(r"\s*[-\u2010-\u2015]+\s*", " - ", text)
        text = re.sub(r"\s*([/(),.:;])\s*", r"\1 ", text)
        text = re.sub(r"\s+", " ", text).strip()
        text = re.sub(r"\s+-\s+-\s+", " - ", text)
        text = re.sub(r"-{2,}", "-", text)
        text = re.sub(r"(\d+)\s*[xX*]\s*(\d+)", lambda m: f"{m.group(1)}x{m.group(2)}", text)

        parts = text.split()
        arabic_prefix: list[str] = []
        rest = parts[:]
        while rest and cls._has_arabic(rest[0]) and not cls._has_latin_or_digit(rest[0]):
            arabic_prefix.append(rest.pop(0))
        if arabic_prefix and any(cls._has_latin_or_digit(part) for part in rest):
            parts = rest + arabic_prefix

        text = " ".join(cls._display_token(part) for part in parts)
        text = re.sub(r"\s+-\s+", " - ", text)
        text = re.sub(r"\s+", " ", text).strip()
        return text or cls.UNKNOWN_PRODUCT_NAME

    @classmethod
    def normalize_product_clean(cls, value: Any) -> str:
        text = cls.normalize_product_display(value)
        text = re.sub(r"[\u064B-\u065F\u0670]", "", text)
        text = re.sub(r"[^\w\u0600-\u06FF\s\-/]", " ", text, flags=re.UNICODE)
        text = re.sub(r"\s*-\s*", " - ", text)
        text = re.sub(r"\s*/\s*", "/", text)
        text = re.sub(r"\s+", " ", text).strip()
        return text.upper() if text else cls.UNKNOWN_PRODUCT_NAME.upper()

    @classmethod
    def product_name_frame(cls, source: Series) -> DataFrame:
        raw = source.astype("string")
        # PERF: deduplicate before applying expensive per-value normalization.
        # For large sale.report extracts many product names repeat thousands of times;
        # computing normalize_product_display / normalize_product_clean once per unique
        # value instead of once per row cuts this step by 80-95%.
        unique_raw = raw.dropna().unique()
        if len(unique_raw) == 0:
            display_map: dict = {}
            clean_map: dict = {}
        else:
            display_map = {v: cls.normalize_product_display(v) for v in unique_raw}
            clean_map = {d: cls.normalize_product_clean(d) for d in display_map.values()}
        display = raw.map(display_map).astype("string")
        # Values that were NA in raw stay NA; fill from map
        display = display.where(raw.notna(), pd.NA)
        clean = display.map(clean_map).astype("string")
        clean = clean.where(display.notna(), pd.NA)
        return pd.DataFrame({"ProductNameRaw": raw, "ProductName": display, "ProductNameClean": clean}, index=source.index)

    @staticmethod
    def clean_product_key_part(series: Series) -> Series:
        """Apply the exact product matching cleanup rules to one key column."""
        source = series.astype("string")
        cleaned = source.apply(lambda value: "" if pd.isna(value) or str(value).strip() == "" else TextUtils.normalize_product_clean(value)).astype("string")
        return cleaned.replace({"NULL": "", "NONE": "", "NAN": "", "<NA>": ""})


def normalize_product_name(value: Any) -> str:
    return TextUtils.normalize_product_display(value)


class ParseUtils:
    @staticmethod
    def year(series: Series) -> Series:
        text = series.astype(str).str.strip()
        year = pd.to_numeric(text, errors="coerce")
        if year.notna().sum() == 0:
            extracted = text.str.extract(r"(\d{4})", expand=False)
            year = pd.to_numeric(extracted, errors="coerce")
        return year.astype("Int64")

    @staticmethod
    def month(series: Series) -> Series:
        text = series.astype(str).str.strip()
        numeric = pd.to_numeric(text, errors="coerce")
        mapping = {
            "jan": 1, "january": 1,
            "feb": 2, "february": 2,
            "mar": 3, "march": 3,
            "apr": 4, "april": 4,
            "may": 5,
            "jun": 6, "june": 6,
            "jul": 7, "july": 7,
            "aug": 8, "august": 8,
            "sep": 9, "sept": 9, "september": 9,
            "oct": 10, "october": 10,
            "nov": 11, "november": 11,
            "dec": 12, "december": 12,
            "fullyear": 0, "full year": 0, "year": 0, "total": 0,
        }
        lower = text.str.lower()
        named = lower.map(mapping)
        short = lower.str[:3].map({
            "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5,
            "jun": 6, "jul": 7, "aug": 8, "sep": 9, "oct": 10,
            "nov": 11, "dec": 12,
        })
        return numeric.fillna(named).fillna(short).astype("Int64")


class DataFrameUtils:
    @staticmethod
    def ensure_columns(df: DataFrame, columns: Sequence[str], default: Any = pd.NA) -> DataFrame:
        out = df.copy()
        for col in columns:
            if col not in out.columns:
                out[col] = default
        return out

    @staticmethod
    def rename_by_alias(df: DataFrame, alias_map: Dict[str, List[str]]) -> DataFrame:
        reverse: Dict[str, str] = {}
        for canonical, aliases in alias_map.items():
            for alias in aliases:
                reverse[re.sub(r"\s+", "", str(alias)).lower()] = canonical

        renamed = {}
        for col in df.columns:
            key = re.sub(r"\s+", "", str(col)).lower()
            renamed[col] = reverse.get(key, col)
        return df.rename(columns=renamed)

    @staticmethod
    def clean_object_columns(df: DataFrame) -> DataFrame:
        out = df.copy()
        for col in out.columns:
            if out[col].dtype == "object":
                out[col] = out[col].apply(TextUtils.norm)
        return out

    @staticmethod
    def latest_attribute(df: DataFrame, grain_cols: List[str], attr_col: str, out_col: str) -> DataFrame:
        if attr_col not in df.columns:
            return pd.DataFrame(columns=grain_cols + [out_col])

        tmp = df[grain_cols + ["order_date_date", attr_col]].copy()
        tmp[attr_col] = tmp[attr_col].apply(TextUtils.norm)
        tmp = tmp[tmp[attr_col].notna()]
        tmp = tmp.dropna(subset=["order_date_date"])
        if tmp.empty:
            return pd.DataFrame(columns=grain_cols + [out_col])

        tmp = (
            tmp.sort_values(grain_cols + ["order_date_date"])
            .drop_duplicates(subset=grain_cols, keep="last")
            .rename(columns={attr_col: out_col})
        )
        return tmp[grain_cols + [out_col]]

    @staticmethod
    def add_key_from_dimension(
        df: DataFrame,
        left_key: str,
        dim: DataFrame,
        dim_key: str,
        new_key: str,
    ) -> DataFrame:
        out = df.copy()

        lookup = (
            dim[[dim_key, new_key]]
            .dropna(subset=[dim_key])
            .drop_duplicates(subset=[dim_key])
        )

        if left_key == dim_key:
            out = out.merge(
                lookup,
                on=left_key,
                how="left",
                validate="m:1",
            )
        else:
            out = out.merge(
                lookup,
                left_on=left_key,
                right_on=dim_key,
                how="left",
                validate="m:1",
            ).drop(columns=[dim_key], errors="ignore")

        return out


class ProductKeyUtils:
    @staticmethod
    def ensure_unique(product_master: DataFrame) -> DataFrame:
        """Make ProductKey a stable row key without removing manual-master rows."""
        out = DataFrameUtils.ensure_columns(product_master.copy(), ["ProductKey"])
        original_keys = out["ProductKey"].astype("string").fillna("").str.strip()
        unique_keys: list[str] = []
        used: set[str] = set()

        for row_number, original_key in enumerate(original_keys, start=1):
            base_key = str(original_key).strip()
            if not base_key:
                candidate = f"PRODUCT|ROW|{row_number:06d}"
            elif base_key not in used:
                candidate = base_key
            else:
                candidate = f"{base_key}|ROW|{row_number:06d}"

            suffix = 2
            while candidate in used:
                candidate = f"{base_key or 'PRODUCT'}|ROW|{row_number:06d}|DUP|{suffix}"
                suffix += 1
            used.add(candidate)
            unique_keys.append(candidate)

        out["ProductKey"] = pd.Series(unique_keys, index=out.index, dtype="string")
        return out


# ============================================================
# Repositories / loaders
# ============================================================
class FileLoader:
    @staticmethod
    def read_any(path: PathLike) -> DataFrame:
        path = Path(path)
        suffix = path.suffix.lower()
        if suffix in {".xlsx", ".xls", ".xlsm"}:
            return pd.read_excel(path)
        if suffix in {".csv", ".txt"}:
            return pd.read_csv(path)
        raise ValueError(f"Unsupported file type: {suffix}")


class SalesLoader:
    def __init__(self, logger: Logger) -> None:
        self.logger = logger

    def load(self, path: PathLike) -> DataFrame:
        raw = FileLoader.read_any(path).copy()
        raw.columns = [str(c).strip() for c in raw.columns]
        existing = [c for c in SALES_COLUMN_MAP if c in raw.columns]
        if not existing:
            raise ValueError(
                "None of the expected sales columns were found.\n"
                f"Expected at least one of: {list(SALES_COLUMN_MAP.keys())}\n"
                f"Got: {list(raw.columns)}"
            )

        df = raw[existing].rename(columns={c: SALES_COLUMN_MAP[c] for c in existing})
        df = df.loc[:, ~df.columns.duplicated()]

        ordered: List[str] = []
        for source_col in SALES_COLUMN_MAP:
            if source_col in raw.columns:
                target_col = SALES_COLUMN_MAP[source_col]
                if target_col in df.columns and target_col not in ordered:
                    ordered.append(target_col)
        df = df[ordered]

        for col in df.select_dtypes(include=["object"]).columns:
            df[col] = df[col].astype(str).str.strip()

        self.logger.ok(f"Loaded sales lines: {len(df):,}")
        return df


class TargetsLoader:
    EXPECTED = [
        "Year", "Month", "YearMonth", "TargetDate",
        "TargetLevel", "Company", "SalesTeam", "SalesTeamKey",
        "Salesperson", "SalesSegment", "City", "Currency",
        "Target_Revenue", "Target_Volume", "ASP_LY", "ASP_ThisYear",
        "DistributionChannel", "ChannelKey", "SalespersonKey",
    ]

    def __init__(self, logger: Logger) -> None:
        self.logger = logger

    def load(self, path: PathLike) -> DataFrame:
        df = pd.read_excel(Path(path)).copy()
        df.columns = [str(c).strip() for c in df.columns]
        df = DataFrameUtils.ensure_columns(df, self.EXPECTED)
        df = df[self.EXPECTED].copy()

        for col in [
            "TargetLevel", "Company", "SalesTeam", "SalesTeamKey", "Salesperson",
            "SalesSegment", "City", "Currency", "DistributionChannel"
        ]:
            df[col] = df[col].apply(TextUtils.norm)

        for col in ["Target_Revenue", "Target_Volume", "ASP_LY", "ASP_ThisYear"]:
            if df[col].dtype == "object":
                df[col] = df[col].astype(str).map(TextUtils.normalize_numeric_string)
            df[col] = pd.to_numeric(df[col], errors="coerce")

        df["Year"] = ParseUtils.year(df["Year"])
        df["Month"] = ParseUtils.month(df["Month"])
        df["TargetDate"] = pd.to_datetime(df["TargetDate"], errors="coerce")

        missing_target_date = df["TargetDate"].isna() & df["Year"].notna() & df["Month"].notna()
        if missing_target_date.any():
            df.loc[missing_target_date, "TargetDate"] = pd.to_datetime(
                dict(
                    year=df.loc[missing_target_date, "Year"].astype(int),
                    month=df.loc[missing_target_date, "Month"].astype(int),
                    day=1,
                ),
                errors="coerce",
            )

        missing_year_month = df["YearMonth"].isna() & df["Year"].notna() & df["Month"].notna()
        if missing_year_month.any():
            df.loc[missing_year_month, "YearMonth"] = (
                df.loc[missing_year_month, "Year"].astype("Int64").astype(str)
                + "-"
                + df.loc[missing_year_month, "Month"].astype("Int64").astype(str).str.zfill(2)
            )

        df = df.dropna(subset=["Year", "Month"]).copy()
        df = df[df["Month"].between(1, 12)]

        self.logger.info(f"Targets rows loaded: {len(df):,}")
        return df.reset_index(drop=True)


class BlockedCustomersLoader:
    EXPECTED = [
        "CustomerID", "CustomerName", "IsBlocked", "BlockedDate",
        "UnblockedDate", "BlockedReason", "Notes",
    ]

    def __init__(self, logger: Logger) -> None:
        self.logger = logger

    def load(self, path: PathLike) -> DataFrame:
        df = pd.read_excel(Path(path)).copy()
        df.columns = [str(c).strip() for c in df.columns]
        df = df.rename(columns={
            "Customer Name": "CustomerName",
            "Customer": "CustomerName",
            "Customer ID": "CustomerID",
            "Customer Code": "CustomerID",
            "Blocked Reason": "BlockedReason",
        })
        df = DataFrameUtils.ensure_columns(df, self.EXPECTED)

        for col in ["CustomerID", "CustomerName", "BlockedReason", "Notes"]:
            df[col] = df[col].apply(TextUtils.norm)

        df["IsBlocked"] = (
            df["IsBlocked"].astype(str).str.strip().str.lower()
            .map({"1": 1, "true": 1, "yes": 1, "y": 1, "0": 0, "false": 0, "no": 0, "n": 0})
            .fillna(0)
            .astype(int)
        )
        df["BlockedDate"] = pd.to_datetime(df["BlockedDate"], errors="coerce")
        df["UnblockedDate"] = pd.to_datetime(df["UnblockedDate"], errors="coerce")
        return df

    def export_template(self, path: PathLike) -> Path:
        out = Path(path)
        out.parent.mkdir(parents=True, exist_ok=True)
        pd.DataFrame(columns=self.EXPECTED).to_excel(out, sheet_name="BlockedCustomers", index=False)
        self.logger.ok(f"Exported blocked customers template: {out}")
        return out


@dataclass
class SalesOrgMaps:
    people_active: DataFrame
    teams: DataFrame
    person_to_teamkey: Dict[str, str]
    teamkey_to_meta: Dict[str, Dict[str, Any]]
    teamname_to_key: Dict[str, str]
    person_to_channel: Dict[str, str]
    teamkey_to_channel: Dict[str, str]


class SalesOrgRepository:
    PEOPLE_ALIAS = {
        "Salesperson": ["Salesperson", "Sales Person", "Name", "Employee", "salesperson", "Sales Man"],
        "TeamKey": ["TeamKey", "Team Key", "Team_Id", "Team Id", "Team Code", "TeamCode", "teamkey", "Key"],
        "Status": ["Status", "STATUS", "status", "ActiveFlag", "Active"],
        "DistributionChannel": ["DistributionChannel", "Distribution Channel", "Channel", "DistChannel", "Distribution"],
    }
    TEAM_ALIAS = {
        "TeamKey": ["TeamKey", "Team Key", "Team_Id", "Team Id", "Team Code", "TeamCode", "Key"],
        "TeamName": ["TeamName", "Team Name", "Name", "sales team", "Team"],
        "Segment": ["Segment", "SalesSegment", "Sales Segment", "Type"],
        "City": ["City", "Branch", "Location"],
        "Company": ["Company", "Org", "Business Unit"],
        "Status": ["Status", "STATUS", "status", "ActiveFlag", "Active"],
    }

    def __init__(self, logger: Logger) -> None:
        self.logger = logger

    def load(self, path: PathLike) -> SalesOrgMaps:
        workbook = ExcelFile(Path(path))
        people_sheet = self._pick_sheet(workbook, PEOPLE_SHEET_CANDIDATES)
        team_sheet = self._pick_sheet(workbook, TEAMS_SHEET_CANDIDATES)
        self.logger.info(f"Using sheets -> People: '{people_sheet}', Teams: '{team_sheet}'")

        people = pd.read_excel(workbook, sheet_name=people_sheet)
        teams = pd.read_excel(workbook, sheet_name=team_sheet)

        people = DataFrameUtils.clean_object_columns(people)
        teams = DataFrameUtils.clean_object_columns(teams)
        people = DataFrameUtils.rename_by_alias(people, self.PEOPLE_ALIAS)
        teams = DataFrameUtils.rename_by_alias(teams, self.TEAM_ALIAS)

        self._validate_people(people)
        self._validate_teams(teams)

        if "DistributionChannel" not in people.columns:
            people["DistributionChannel"] = pd.NA

        people_active = people.copy()
        if "Status" in people_active.columns:
            people_active = people_active[
                people_active["Status"].astype(str).str.upper().eq("ACTIVE")
            ].copy()

        people_active["TeamKey"] = people_active["TeamKey"].astype(str)
        teams["TeamKey"] = teams["TeamKey"].astype(str)

        person_to_teamkey = (
            people_active.dropna(subset=["Salesperson", "TeamKey"])
            .set_index("Salesperson")["TeamKey"]
            .to_dict()
        )
        person_to_channel = (
            people_active.dropna(subset=["Salesperson"])
            .set_index("Salesperson")["DistributionChannel"]
            .to_dict()
        )
        teamname_to_key = {
            str(row["TeamName"]): str(row["TeamKey"])
            for _, row in teams.dropna(subset=["TeamKey", "TeamName"]).iterrows()
        }
        teamkey_to_meta = {
            str(row["TeamKey"]): {
                "team_key": row.get("TeamKey"),
                "team_name": row.get("TeamName"),
                "segment": row.get("Segment"),
                "city": row.get("City"),
                "company": row.get("Company"),
                "status": row.get("Status") if "Status" in row.index else pd.NA,
            }
            for _, row in teams.dropna(subset=["TeamKey"]).iterrows()
        }
        teamkey_to_channel = self._build_teamkey_to_channel(people_active)

        return SalesOrgMaps(
            people_active=people_active,
            teams=teams,
            person_to_teamkey=person_to_teamkey,
            teamkey_to_meta=teamkey_to_meta,
            teamname_to_key=teamname_to_key,
            person_to_channel=person_to_channel,
            teamkey_to_channel=teamkey_to_channel,
        )

    @staticmethod
    def _pick_sheet(workbook: ExcelFile, candidates: List[str]) -> str:
        available = workbook.sheet_names
        lower_map = {name.lower(): name for name in available}
        for candidate in candidates:
            if candidate.lower() in lower_map:
                return lower_map[candidate.lower()]
        raise ValueError(f"Expected one of {candidates}, available sheets: {available}")

    @staticmethod
    def _validate_people(df: DataFrame) -> None:
        for required in ["Salesperson", "TeamKey"]:
            if required not in df.columns:
                raise KeyError(f"[People] Missing required column '{required}'. Columns: {list(df.columns)}")

    @staticmethod
    def _validate_teams(df: DataFrame) -> None:
        for required in ["TeamKey", "TeamName", "Segment", "City", "Company"]:
            if required not in df.columns:
                raise KeyError(f"[Teams] Missing required column '{required}'. Columns: {list(df.columns)}")

    @staticmethod
    def _build_teamkey_to_channel(people_active: DataFrame) -> Dict[str, str]:
        result: Dict[str, str] = {}
        tmp = people_active[["TeamKey", "DistributionChannel"]].dropna(subset=["TeamKey"])
        for team_key, group in tmp.groupby("TeamKey"):
            values = group["DistributionChannel"].dropna().astype(str).str.strip()
            values = values[values != ""]
            if values.empty:
                result[str(team_key)] = "Unknown"
                continue
            unique = sorted(set(values.tolist()))
            if len(unique) == 1:
                result[str(team_key)] = unique[0]
            else:
                mode = values.mode()
                result[str(team_key)] = str(mode.iloc[0]) if len(mode) == 1 else "Mixed"
        return result


class ProductMasterLoader:
    EXPECTED = [
        "Company", "Category", "Brand", "SubBrand", "Family",
        "ProductName", "OdooProductName", "ProductLevel", "SKU", "Size", "IsActive", "ProductKey",
    ]

    def __init__(self, logger: Logger, export_conflicts: bool, base_dir: Path) -> None:
        self.logger = logger
        self.export_conflicts = export_conflicts
        self.base_dir = base_dir

    def load(self, path: PathLike) -> DataFrame:
        df = pd.read_excel(Path(path)).copy()
        df = df.rename(columns={
            "Odoo Nmae": "OdooProductName",
            "Odoo Name": "OdooProductName",
            "OdooProduct Name": "OdooProductName",
            "Subbrand": "SubBrand",
            "Sub Brand": "SubBrand",
            "ProductLevet": "ProductLevel",
            "Product Level": "ProductLevel",
        })
        df = DataFrameUtils.ensure_columns(df, self.EXPECTED)
        df = self._clean(df)
        valid_product = df["ProductName"].astype("string").str.strip().replace("", pd.NA).notna()
        dropped = int((~valid_product).sum())
        if dropped:
            self.logger.warn(f"Dropped {dropped} PRODUCTS.xlsx rows without ProductName")
        df = df.loc[valid_product].reset_index(drop=True)
        df = ProductKeyUtils.ensure_unique(df)
        return df

    def _clean(self, df: DataFrame) -> DataFrame:
        out = df.copy()
        for col in self.EXPECTED:
            if col in out.columns:
                out[col] = (
                    out[col].astype("string")
                    .str.replace(r"[\r\n\t]+", " ", regex=True)
                    .str.strip()
                    .replace({"": pd.NA, "nan": pd.NA, "None": pd.NA, "-": pd.NA, "—": pd.NA})
                )
        out["Company"] = out["Company"].apply(TextUtils.normalize_company)
        out["ProductLevel"] = out["ProductLevel"].astype("string").str.strip().str.upper()
        out.loc[out["ProductLevel"].isna() | (out["ProductLevel"] == ""), "ProductLevel"] = pd.NA
        out["IsActive"] = self._coerce_isactive(out["IsActive"])
        return out

    @staticmethod
    def _coerce_isactive(series: Series, default: int = 1) -> Series:
        mapped = (
            series.astype(str).str.strip().str.lower()
            .map({
                "1": 1, "true": 1, "yes": 1, "y": 1, "active": 1,
                "0": 0, "false": 0, "no": 0, "n": 0, "inactive": 0,
            })
        )
        return mapped.fillna(default).astype(int)

# ============================================================
# Transformers
# ============================================================
class SalesCleaner:
    def clean_order_numbers(self, df: DataFrame) -> DataFrame:
        if "order_number" not in df.columns:
            return df
        out = df.copy()
        out["order_number"] = (
            out["order_number"].astype(str).str.strip()
            .replace({"": pd.NA, "nan": pd.NA, "NaN": pd.NA, "None": pd.NA})
        )
        return out

    def clean_customer(self, df: DataFrame) -> DataFrame:
        out = df.copy()
        if "customer" in out.columns:
            out["customer_raw"] = out["customer"]
            out["customer"] = (
                out["customer"].astype(str)
                .str.replace(r"[\r\n\t]+", " ", regex=True)
                .str.strip()
                .replace({"": pd.NA, "nan": pd.NA, "NaN": pd.NA, "None": pd.NA})
                .apply(TextUtils.norm)
            )
        if "customer_id" in out.columns:
            out["customer_id_raw"] = out["customer_id"]
            out["customer_id"] = (
                out["customer_id"].astype(str)
                .str.replace(r"[\r\n\t]+", " ", regex=True)
                .str.strip()
                .replace({"": pd.NA, "nan": pd.NA, "NaN": pd.NA, "None": pd.NA})
                .apply(TextUtils.norm)
            )
        return out

    def clean_product_names(self, df: DataFrame) -> DataFrame:
        if "product_name" not in df.columns:
            return df
        out = df.copy()
        out["product_name_raw"] = out["product_name"]
        out["OdooProductName"] = out["product_name_raw"]
        names = TextUtils.product_name_frame(out["product_name_raw"])
        out["product_name"] = names["ProductName"]
        out["product_name_clean"] = names["ProductNameClean"]
        out["ProductNameRaw"] = names["ProductNameRaw"]
        out["ProductName"] = names["ProductName"]
        out["ProductNameClean"] = names["ProductNameClean"]
        out["OdooProductNameClean"] = names["ProductNameClean"]
        # PERF: Replaced apply(lambda) with vectorized str.contains (C-level regex engine)
        out["is_discount"] = out["product_name_raw"].astype(str).str.contains(
            r"discount|تخفيض", case=False, na=False, regex=True
        )
        out.loc[out["is_discount"], "product_name"] = "Discount"
        out.loc[out["is_discount"], "product_name_clean"] = "DISCOUNT"
        out.loc[out["is_discount"], "ProductName"] = "Discount"
        out.loc[out["is_discount"], "ProductNameClean"] = "DISCOUNT"
        out.loc[out["is_discount"], "OdooProductNameClean"] = "DISCOUNT"
        out["product_name_norm"] = out["product_name_clean"]
        return out

    def clean_quantity(self, df: DataFrame) -> DataFrame:
        if "quantity" not in df.columns:
            return df
        out = df.copy()
        if out["quantity"].dtype == "object":
            out["quantity"] = out["quantity"].astype(str).str.replace(",", ".", regex=False).str.strip()
        out["quantity"] = pd.to_numeric(out["quantity"], errors="coerce")
        out["Volume"] = out["quantity"]
        return out

    def clean_value(self, df: DataFrame) -> DataFrame:
        if "line_total" not in df.columns:
            return df
        out = df.copy()
        if out["line_total"].dtype == "object":
            out["line_total"] = out["line_total"].astype(str).map(TextUtils.normalize_numeric_string)
        out["line_total"] = pd.to_numeric(out["line_total"], errors="coerce")
        out["Value"] = out["line_total"]
        return out

    def clean_order_dates(self, df: DataFrame) -> DataFrame:
        if "order_date" not in df.columns:
            return df
        out = df.copy()
        out["order_date_raw"] = out["order_date"]
        out["order_date"] = (
            out["order_date"].astype(str).str.strip()
            .replace({"": pd.NA, "nan": pd.NA, "NaN": pd.NA, "None": pd.NA})
        )
        out["order_date"] = pd.to_datetime(out["order_date"], errors="coerce")
        # Reporting date for Power BI. Odoo UTC datetimes must already be
        # converted to local business time before this point, then normalized.
        # Do not timezone-shift this date-only value after it has been created.
        out["order_date_date"] = out["order_date"].dt.normalize()
        return out

    def filter_dashboard_rows(self, df: DataFrame) -> DataFrame:
        out = df.copy()
        if "invoice_status" in out.columns:
            out["invoice_status"] = out["invoice_status"].astype(str).str.strip()
            out = out[out["invoice_status"].str.lower() != "no"]
        return out

    def clean_salesperson(self, df: DataFrame) -> DataFrame:
        if "salesperson" not in df.columns:
            return df
        out = df.copy()
        out["salesperson_raw"] = out["salesperson"]
        out["salesperson"] = (
            out["salesperson"].astype(str)
            .str.replace(r"[\r\n\t]+", " ", regex=True)
            .str.strip()
            .replace({"": pd.NA, "nan": pd.NA, "NaN": pd.NA, "None": pd.NA})
        )
        return out

    def clean_company(self, df: DataFrame) -> DataFrame:
        if "company" not in df.columns:
            return df
        out = df.copy()
        out["company_raw"] = out["company"]
        out["company"] = (
            out["company"].astype(str).str.strip()
            .replace({"": pd.NA, "nan": pd.NA, "NaN": pd.NA, "None": pd.NA})
        )
        out["Company"] = out["company"].apply(TextUtils.normalize_company)
        return out

    def add_first_purchase_date(self, df: DataFrame) -> DataFrame:
        required = ["customer", "order_date_date"]
        missing = [c for c in required if c not in df.columns]
        if missing:
            raise KeyError(f"Missing columns for first purchase logic: {missing}")

        out = df.copy()
        grain_cols = ["customer"]
        if "customer_id" in out.columns and out["customer_id"].notna().any():
            grain_cols = ["customer_id", "customer"]

        first_purchase = (
            out.loc[out["customer"].notna() & out["order_date_date"].notna(), grain_cols + ["order_date_date"]]
            .groupby(grain_cols, as_index=False)["order_date_date"]
            .min()
            .rename(columns={"order_date_date": "First_Purchase_Date"})
        )
        first_purchase["First_Purchase_Date"] = pd.to_datetime(first_purchase["First_Purchase_Date"])
        return out.merge(first_purchase, on=grain_cols, how="left", validate="m:1")

    def add_invoice_summary(self, df: DataFrame) -> DataFrame:
        required = ["order_number", "Value", "invoice_status", "order_date_date"]
        missing = [c for c in required if c not in df.columns]
        if missing:
            raise KeyError(f"Missing columns for invoice classification: {missing}")

        out = df.copy()
        base = out[
            out["order_number"].notna()
            & out["Value"].notna()
            & out["invoice_status"].astype(str).str.strip().str.lower().eq("invoiced")
        ].copy()

        if base.empty:
            out["InvoiceValue"] = pd.NA
            out["InvoiceClass"] = pd.NA
            out["InvoiceYear"] = pd.NA
            return out

        invoice_summary = (
            base.groupby("order_number", as_index=False)
            .agg(
                OrderDate=("order_date_date", "max"),
                InvoiceValue=("Value", "sum"),
            )
        )
        invoice_summary["InvoiceYear"] = invoice_summary["OrderDate"].dt.year.astype("Int64")
        invoice_summary["InvoiceClass"] = invoice_summary["InvoiceValue"].apply(self._classify_invoice)

        out = out.merge(
            invoice_summary[["order_number", "InvoiceValue", "InvoiceClass", "InvoiceYear"]],
            on="order_number",
            how="left",
            validate="m:1",
        )
        out["InvoiceValue"] = pd.to_numeric(out["InvoiceValue"], errors="coerce")
        return out

    @staticmethod
    def _classify_invoice(value: Any) -> Any:
        if pd.isna(value):
            return pd.NA
        if value > 50000:
            return "Class A (>50K)"
        if 25000 <= value <= 50000:
            return "Class B (25K - 50K)"
        if 5000 <= value < 25000:
            return "Class C (5K - 25K)"
        return "Class D (<5K)"


class ProductMapper:
    RAW_KEY_PREFIX = "RAW_FROM_ODOO"

    @staticmethod
    def _coalesce_text(df: DataFrame, names: Sequence[str]) -> Series:
        result = pd.Series(pd.NA, index=df.index, dtype="string")
        for name in names:
            if name not in df.columns:
                continue
            candidate = df[name].astype("string").str.strip().replace("", pd.NA)
            result = result.fillna(candidate)
        return result

    @classmethod
    def _raw_product_key(cls, company: Any, clean_name: Any) -> str:
        company_part = TextUtils.clean_product_key_part(pd.Series([company])).iloc[0] or "UNKNOWN"
        name_part = TextUtils.clean_product_key_part(pd.Series([clean_name])).iloc[0] or "UNKNOWN_PRODUCT"
        digest = hashlib.sha1(f"{company_part}|{name_part}".encode("utf-8")).hexdigest()[:12].upper()
        return f"{cls.RAW_KEY_PREFIX}|{digest}"

    @staticmethod
    def attach(df_sales: DataFrame, product_master: DataFrame, odoo_products: Optional[DataFrame] = None) -> DataFrame:
        if "product_name_clean" not in df_sales.columns:
            raise KeyError("product_name_clean missing. Run clean_product_names() first.")

        out = df_sales.copy()
        raw_odoo_name = ProductMapper._coalesce_text(
            out,
            ["OdooProductName", "product_name_raw", "ProductNameRaw", "product_name", "ProductName", "product_name_clean"],
        )
        odoo_names = TextUtils.product_name_frame(raw_odoo_name)
        out["OdooProductName"] = odoo_names["ProductNameRaw"]
        out["OdooProductNameClean"] = odoo_names["ProductNameClean"]
        out["_odoo_product_name_key"] = TextUtils.clean_product_key_part(out["OdooProductNameClean"])

        pm = ProductDimensionBuilder().build(product_master, raw_odoo_products=odoo_products)
        pm["IsActive"] = pm["IsActive"].fillna(1).astype(int)
        bridge_columns = [
            "ProductKey", "ProductName", "ProductNameClean", "OdooProductID", "OdooProductName",
            "OdooProductNameClean", "SKU", "Size", "ProductLevel", "Company",
        ]
        pm = DataFrameUtils.ensure_columns(pm, bridge_columns)
        out["OdooProductID"] = pd.to_numeric(out.get("OdooProductID", pd.Series(pd.NA, index=out.index)), errors="coerce").astype("Int64")
        pm["OdooProductID"] = pd.to_numeric(pm["OdooProductID"], errors="coerce").astype("Int64")
        pm["_odoo_product_name_key"] = TextUtils.clean_product_key_part(pm["OdooProductNameClean"])
        pm["_odoo_product_id_key"] = pm["OdooProductID"].astype("string").fillna("")
        out["_odoo_product_id_key"] = out["OdooProductID"].astype("string").fillna("")

        def make_bridge(key: str) -> DataFrame:
            bridge = pm[pm[key].astype("string").str.strip().ne("")].copy()
            bridge = (
                bridge[bridge_columns + [key, "IsActive"]]
                .sort_values([key, "IsActive", "ProductKey"], ascending=[True, False, True], na_position="last")
                .drop_duplicates(subset=[key], keep="first")
            )
            return bridge.rename(columns={column: f"_mapped_{column}" for column in bridge_columns})

        primary = make_bridge("_odoo_product_id_key")
        out = out.merge(
            primary.drop(columns=["IsActive"], errors="ignore"),
            on="_odoo_product_id_key",
            how="left",
            validate="m:1",
        )
        out["ProductMappingStatus"] = pd.Series(pd.NA, index=out.index, dtype="string")
        primary_matched = out["_mapped_ProductKey"].notna()
        out.loc[primary_matched, "ProductMappingStatus"] = "MatchedByOdooProductID"

        name_lookup = make_bridge("_odoo_product_name_key")
        need_name_match = out["_mapped_ProductKey"].isna()
        if need_name_match.any():
            name_candidate = out.loc[need_name_match, ["_odoo_product_name_key"]].merge(
                name_lookup.drop(columns=["IsActive"], errors="ignore"),
                on="_odoo_product_name_key",
                how="left",
                validate="m:1",
            )
            name_candidate.index = out.index[need_name_match]
            name_hit = name_candidate["_mapped_ProductKey"].notna()
            if name_hit.any():
                hit_index = name_candidate.index[name_hit]
                for column in bridge_columns:
                    out.loc[hit_index, f"_mapped_{column}"] = name_candidate.loc[name_hit, f"_mapped_{column}"].to_numpy()
                out.loc[hit_index, "ProductMappingStatus"] = "MatchedByOdooProductName"

        matched = out["_mapped_ProductKey"].notna()
        raw_display = out["OdooProductName"].apply(TextUtils.normalize_product_display)
        raw_clean = raw_display.apply(TextUtils.normalize_product_clean)
        company_source = out["company"] if "company" in out.columns else out.get("Company", pd.Series(pd.NA, index=out.index))
        raw_keys = [
            ProductMapper._raw_product_key(company, clean_name)
            for company, clean_name in zip(company_source, raw_clean)
        ]

        out["ProductKey"] = out["_mapped_ProductKey"].where(matched, pd.Series(raw_keys, index=out.index, dtype="string"))
        out["ProductName"] = out["_mapped_ProductName"].where(matched, raw_display)
        out["ProductNameClean"] = out["_mapped_ProductNameClean"].where(matched, raw_clean)
        out["product_name"] = out["ProductName"]
        out["product_name_clean"] = out["ProductNameClean"]
        out["ProductNameRaw"] = out["ProductName"]
        out["SKU"] = out["_mapped_SKU"].where(matched, pd.NA)
        out["Size"] = out["_mapped_Size"].where(matched, pd.NA)
        out["ProductLevel"] = out["_mapped_ProductLevel"].where(matched, pd.NA)
        out["ProductMappingStatus"] = out["ProductMappingStatus"].fillna("RawFromOdoo")
        out["ProductMappingReason"] = np.select(
            [
                out["ProductMappingStatus"].eq("MatchedByOdooProductID"),
                out["ProductMappingStatus"].eq("MatchedByOdooProductName"),
            ],
            [
                "Matched Odoo product_id to Dim_Product[OdooProductID]",
                "Matched Odoo normalized product name to Dim_Product[OdooProductNameClean]",
            ],
            default="Loaded directly from Odoo product data",
        )
        out["ProductJoinKey"] = out["_mapped_SKU"].where(matched, out["ProductKey"]).astype("string")
        return out.drop(
            columns=[
                "_odoo_product_id_key",
                "_odoo_product_name_key",
                *[f"_mapped_{column}" for column in bridge_columns],
            ],
            errors="ignore",
        )

class SalesOrgEnricher:
    def __init__(self, sales_org: SalesOrgMaps) -> None:
        self.sales_org = sales_org

    def enrich_sales(self, df: DataFrame) -> DataFrame:
        if "sales_team" not in df.columns or "salesperson" not in df.columns:
            raise KeyError("Columns 'sales_team' and 'salesperson' are required.")

        out = df.copy()
        out["sales_team_raw"] = out["sales_team"]
        out["sales_team_norm"] = out["sales_team"].apply(TextUtils.norm)
        out["salesperson_norm"] = out["salesperson"].apply(TextUtils.norm)

        out["SalesTeamKey"] = out["sales_team_norm"].map(self.sales_org.teamname_to_key)
        need_fallback = out["SalesTeamKey"].isna() & out["salesperson_norm"].notna()
        out.loc[need_fallback, "SalesTeamKey"] = out.loc[need_fallback, "salesperson_norm"].map(
            self.sales_org.person_to_teamkey
        )

        for col in ["SalesTeam", "SalesSegment", "SalesTeamCompany", "SalesCity", "SalesTeamStatus"]:
            out[col] = pd.NA

        # PERF: Replaced apply(fill_meta, axis=1) with vectorized merge.
        # Old approach iterated every sales row in Python; new approach is O(1) dict → DataFrame join.
        if self.sales_org.teamkey_to_meta:
            _meta_rows = [
                {
                    "_meta_SalesTeamKey": k,
                    "SalesTeam": v.get("team_name"),
                    "SalesSegment": v.get("segment"),
                    "SalesTeamCompany": v.get("company"),
                    "SalesCity": v.get("city"),
                    "SalesTeamStatus": v.get("status"),
                }
                for k, v in self.sales_org.teamkey_to_meta.items()
            ]
            _meta_df = pd.DataFrame(_meta_rows)
            out = out.drop(columns=["SalesTeam", "SalesSegment", "SalesTeamCompany", "SalesCity", "SalesTeamStatus"], errors="ignore")
            _stk_col = out["SalesTeamKey"].astype("string")
            out = out.merge(
                _meta_df.rename(columns={"_meta_SalesTeamKey": "SalesTeamKey"}),
                on="SalesTeamKey",
                how="left",
            )
            for _col in ["SalesTeam", "SalesSegment", "SalesTeamCompany", "SalesCity", "SalesTeamStatus"]:
                if _col not in out.columns:
                    out[_col] = pd.NA

        # clean names
        out["sales_team"] = out["SalesTeam"]

        # legacy compatibility columns for Power BI
        out["sales_team_key"] = out["SalesTeamKey"]
        out["sales_team_segment"] = out["SalesSegment"]
        out["sales_team_city"] = out["SalesCity"]
        out["sales_team_status"] = out["SalesTeamStatus"]

        return out.drop(columns=["sales_team_norm", "salesperson_norm"], errors="ignore")

    def normalize_targets(self, df_targets: DataFrame) -> DataFrame:
        out = df_targets.copy()
        out = DataFrameUtils.ensure_columns(out, ["DistributionChannel", "SalespersonKey", "ChannelKey", "TargetLevel"])

        missing_level = out["TargetLevel"].isna()
        out.loc[missing_level & out["Salesperson"].notna(), "TargetLevel"] = "Salesperson"
        out.loc[missing_level & out["Salesperson"].isna(), "TargetLevel"] = "SalesTeam"

        out["SalesTeamKey"] = out["SalesTeamKey"].astype("string")

        need_team_key = out["SalesTeamKey"].isna() & out["SalesTeam"].notna()
        out.loc[need_team_key, "SalesTeamKey"] = out.loc[need_team_key, "SalesTeam"].map(self.sales_org.teamname_to_key)

        need_person_key = out["SalesTeamKey"].isna() & out["Salesperson"].notna()
        out.loc[need_person_key, "SalesTeamKey"] = out.loc[need_person_key, "Salesperson"].map(self.sales_org.person_to_teamkey)

        # PERF: Replaced apply(fill_from_teamkey, axis=1) with vectorized conditional merge.
        # Old approach iterated every targets row in Python.
        if self.sales_org.teamkey_to_meta:
            _meta_rows = [
                {
                    "SalesTeamKey": k,
                    "_m_SalesTeam": v.get("team_name"),
                    "_m_SalesSegment": v.get("segment"),
                    "_m_City": v.get("city"),
                    "_m_Company": v.get("company"),
                }
                for k, v in self.sales_org.teamkey_to_meta.items()
            ]
            _meta_df = pd.DataFrame(_meta_rows)
            out = out.merge(_meta_df, on="SalesTeamKey", how="left")
            for _col, _meta_col in [
                ("SalesTeam", "_m_SalesTeam"),
                ("SalesSegment", "_m_SalesSegment"),
                ("City", "_m_City"),
                ("Company", "_m_Company"),
            ]:
                if _meta_col in out.columns:
                    if _col not in out.columns:
                        out[_col] = pd.NA
                    # Only fill rows where the existing value is null (mirrors original logic)
                    out[_col] = out[_col].where(out[_col].notna(), out[_meta_col])
                    out = out.drop(columns=[_meta_col])

        need_channel = out["DistributionChannel"].isna() | (out["DistributionChannel"].astype(str).str.strip() == "")
        mask_salesperson = need_channel & out["Salesperson"].notna()
        out.loc[mask_salesperson, "DistributionChannel"] = out.loc[mask_salesperson, "Salesperson"].map(self.sales_org.person_to_channel)

        mask_team = (
            (out["DistributionChannel"].isna() | (out["DistributionChannel"].astype(str).str.strip() == ""))
            & out["SalesTeamKey"].notna()
        )
        out.loc[mask_team, "DistributionChannel"] = out.loc[mask_team, "SalesTeamKey"].astype(str).map(self.sales_org.teamkey_to_channel)
        out["DistributionChannel"] = out["DistributionChannel"].fillna("Unknown").astype(str)
        return out


# ============================================================
# Dimensions
# ============================================================
class DateDimensionBuilder:
    def __init__(self, weekly_rest_day_name: str) -> None:
        self.weekly_rest_day_name = weekly_rest_day_name

    def build(self, df_sales: DataFrame) -> DataFrame:
        if "order_date_date" not in df_sales.columns:
            raise KeyError("order_date_date column not found.")

        dates = pd.to_datetime(df_sales["order_date_date"], errors="coerce").dropna()
        if dates.empty:
            raise ValueError("order_date_date has no valid values.")

        start = dates.min().normalize()
        end = dates.max().normalize()
        dim = pd.DataFrame({"Date": pd.date_range(start=start, end=end, freq="D")})
        dim["Year"] = dim["Date"].dt.year
        dim["Month"] = dim["Date"].dt.month
        dim["MonthName"] = dim["Date"].dt.strftime("%b")
        dim["YearMonth"] = dim["Date"].dt.to_period("M").astype(str)
        dim["Quarter"] = dim["Date"].dt.to_period("Q").astype(str)
        dim["DayOfMonth"] = dim["Date"].dt.day
        dim["DayOfYear"] = dim["Date"].dt.dayofyear
        dim["WeekdayNumber"] = dim["Date"].dt.weekday
        dim["WeekdayName"] = dim["Date"].dt.strftime("%A")
        dim["IsWeeklyRestDay"] = (dim["WeekdayName"] == self.weekly_rest_day_name).astype(int)
        dim.insert(0, "DateKey", dim["Date"].dt.strftime("%Y%m%d").astype(int))
        return dim


class DistributionChannelDimensionBuilder:
    def build(self, people_active: DataFrame, targets: DataFrame) -> DataFrame:
        people_channels = pd.Series([], dtype="string")
        if "DistributionChannel" in people_active.columns:
            people_channels = people_active["DistributionChannel"].astype("string")

        target_channels = pd.Series([], dtype="string")
        if "DistributionChannel" in targets.columns:
            target_channels = targets["DistributionChannel"].astype("string")

        all_values = pd.concat([people_channels, target_channels], ignore_index=True)
        all_values = all_values.dropna().astype(str).str.strip()
        all_values = all_values[all_values != ""]

        unique = sorted(set(all_values.tolist()))
        if "Unknown" not in unique:
            unique = ["Unknown"] + unique

        dim = pd.DataFrame({"DistributionChannel": unique})
        dim.insert(0, "ChannelKey", range(1, len(dim) + 1))
        return dim


class SalespersonDimensionBuilder:
    def __init__(self, sales_org: SalesOrgMaps, settings: PipelineSettings) -> None:
        self.sales_org = sales_org
        self.settings = settings

    def build(self, df_sales: DataFrame) -> DataFrame:
        if "salesperson" not in df_sales.columns:
            raise KeyError("salesperson column not found.")

        dim = (
            df_sales[["salesperson"]].dropna().drop_duplicates()
            .sort_values("salesperson")
            .reset_index(drop=True)
        )
        dim.insert(0, "SalespersonKey", dim.index + 1)
        dim["SalesTeamKey"] = dim["salesperson"].apply(TextUtils.norm).map(self.sales_org.person_to_teamkey)
        dim["SalesTeamKey"] = dim["SalesTeamKey"].astype("string").fillna(self.settings.unknown_team_key).astype(str)
        dim["DistributionChannel"] = dim["salesperson"].apply(TextUtils.norm).map(self.sales_org.person_to_channel)
        dim["DistributionChannel"] = dim["DistributionChannel"].astype("string").fillna(self.settings.unknown_channel_name).astype(str)

        if not (dim["SalespersonKey"] == self.settings.unknown_salesperson_key).any():
            dim = pd.concat([
                pd.DataFrame([{
                    "SalespersonKey": self.settings.unknown_salesperson_key,
                    "salesperson": "Unknown_Salesperson",
                    "SalesTeamKey": self.settings.unknown_team_key,
                    "DistributionChannel": self.settings.unknown_channel_name,
                }]),
                dim,
            ], ignore_index=True)

        return dim.reset_index(drop=True)


class SalesTeamDimensionBuilder:
    def __init__(self, sales_org: SalesOrgMaps, settings: PipelineSettings) -> None:
        self.sales_org = sales_org
        self.settings = settings

    def build(self) -> DataFrame:
        dim = self.sales_org.teams.copy()
        dim = dim.rename(columns={
            "TeamKey": "SalesTeamKey",
            "TeamName": "SalesTeam",
            "Segment": "SalesSegment",
            "City": "SalesCity",
            "Company": "SalesTeamCompany",
            "Status": "SalesTeamStatus",
        })
        keep = ["SalesTeamKey", "SalesTeam", "SalesSegment", "SalesCity", "SalesTeamCompany", "SalesTeamStatus"]
        dim = dim[[col for col in keep if col in dim.columns]].drop_duplicates()
        dim["SalesTeamKey"] = dim["SalesTeamKey"].astype(str)

        dim["SalesTeam"] = dim["SalesTeam"].fillna("Unknown_Team")
        for col in ["SalesSegment", "SalesCity", "SalesTeamCompany"]:
            if col in dim.columns:
                dim[col] = dim[col].fillna("Unknown")
        if "SalesTeamStatus" in dim.columns:
            dim["SalesTeamStatus"] = dim["SalesTeamStatus"].fillna("ACTIVE")

        if not (dim["SalesTeamKey"] == self.settings.unknown_team_key).any():
            dim = pd.concat([
                dim,
                pd.DataFrame([{
                    "SalesTeamKey": self.settings.unknown_team_key,
                    "SalesTeam": "Unknown_Team",
                    "SalesSegment": "Unknown",
                    "SalesCity": "Unknown",
                    "SalesTeamCompany": "Unknown",
                    "SalesTeamStatus": "ACTIVE",
                }]),
            ], ignore_index=True)

        return dim.sort_values(["SalesTeamCompany", "SalesCity", "SalesTeam"], na_position="last").reset_index(drop=True)


class CompanyDimensionBuilder:
    def build(self, df_sales: DataFrame) -> DataFrame:
        if "Company" not in df_sales.columns:
            raise KeyError("Company column not found.")
        dim = (
            df_sales[["Company"]].dropna().drop_duplicates()
            .sort_values("Company")
            .reset_index(drop=True)
        )
        dim.insert(0, "CompanyKey", dim.index + 1)
        return dim


class ProductDimensionBuilder:
    PRODUCT_SOURCE_INPUT = "Input Master"
    PRODUCT_SOURCE_ODOO = "Odoo"
    PRODUCT_SOURCE_UNMAPPED = PRODUCT_SOURCE_ODOO
    COLUMNS = [
        "ProductKey", "Company", "Category", "Brand", "SubBrand", "Family",
        "OdooProductID", "OdooProductTemplateID",
        "ProductNameRaw", "ProductName", "ProductNameClean", "OdooProductName",
        "OdooProductNameClean", "ProductLevel", "SKU", "Size", "IsActive",
        "ProductMappingStatus", "ProductMappingReason", "ProductSource",
        "IsMappedProduct",
    ]

    def build(
        self,
        product_master: DataFrame,
        raw_sales: Optional[DataFrame] = None,
        raw_odoo_products: Optional[DataFrame] = None,
    ) -> DataFrame:
        # Preserve manual product rows for enrichment, then append every product
        # received from Odoo so the product dimension is not gated by PRODUCTS.xlsx.
        pm = product_master.copy()
        pm = DataFrameUtils.ensure_columns(pm, ["ProductName"])
        valid_product = pm["ProductName"].astype("string").str.strip().replace("", pd.NA).notna()
        pm = ProductKeyUtils.ensure_unique(pm.loc[valid_product].reset_index(drop=True))
        pm = DataFrameUtils.ensure_columns(pm, ["OdooProductName", "ProductName", *self.COLUMNS])
        names = TextUtils.product_name_frame(pm["ProductName"])
        pm["ProductNameRaw"] = names["ProductNameRaw"]
        pm["ProductName"] = names["ProductName"]
        pm["ProductNameClean"] = names["ProductNameClean"]
        odoo_names = TextUtils.product_name_frame(pm["OdooProductName"])
        pm["OdooProductNameClean"] = odoo_names["ProductNameClean"].where(
            pm["OdooProductName"].astype("string").str.strip().replace("", pd.NA).notna(),
            pd.NA,
        )
        pm["OdooProductID"] = pd.to_numeric(pm["OdooProductID"], errors="coerce").astype("Int64")
        pm["OdooProductTemplateID"] = pd.to_numeric(pm["OdooProductTemplateID"], errors="coerce").astype("Int64")
        pm["ProductMappingStatus"] = pm["ProductMappingStatus"].fillna("OfficialProductMaster")
        pm["ProductMappingReason"] = pm["ProductMappingReason"].fillna("Loaded from PRODUCTS.xlsx")
        pm["ProductSource"] = self.PRODUCT_SOURCE_INPUT
        pm["IsMappedProduct"] = True
        out = pm[self.COLUMNS].copy()

        odoo_catalog = self._odoo_catalog_rows(raw_odoo_products)
        if not odoo_catalog.empty:
            existing_ids = set(pd.to_numeric(out["OdooProductID"], errors="coerce").dropna().astype("Int64").astype("string"))
            existing_names = set(TextUtils.clean_product_key_part(out["OdooProductNameClean"]))
            catalog = odoo_catalog[
                ~odoo_catalog["OdooProductID"].astype("string").isin(existing_ids)
                & ~TextUtils.clean_product_key_part(odoo_catalog["OdooProductNameClean"]).isin(existing_names)
            ].copy()
            if not catalog.empty:
                out = pd.concat([out, catalog[self.COLUMNS]], ignore_index=True)

        if raw_sales is not None and not raw_sales.empty and "ProductMappingStatus" in raw_sales.columns:
            raw = raw_sales[raw_sales["ProductMappingStatus"].astype("string").eq("RawFromOdoo")].copy()
            if not raw.empty:
                raw = DataFrameUtils.ensure_columns(raw, self.COLUMNS)
                company_fallback = (
                    raw["company"].apply(TextUtils.normalize_company)
                    if "company" in raw.columns
                    else raw["company_final"].apply(TextUtils.normalize_company)
                    if "company_final" in raw.columns
                    else pd.NA
                )
                raw["Company"] = raw["Company"].where(
                    raw["Company"].astype("string").str.strip().replace("", pd.NA).notna(),
                    company_fallback,
                )
                raw["ProductNameRaw"] = raw["ProductName"].where(
                    raw["ProductName"].astype("string").str.strip().replace("", pd.NA).notna(),
                    raw["OdooProductName"],
                )
                raw["OdooProductName"] = raw["OdooProductName"].where(
                    raw["OdooProductName"].astype("string").str.strip().replace("", pd.NA).notna(),
                    raw["ProductNameRaw"],
                )
                raw["ProductName"] = raw["OdooProductName"].apply(TextUtils.normalize_product_display)
                raw["ProductNameClean"] = raw["ProductName"].apply(TextUtils.normalize_product_clean)
                raw["OdooProductNameClean"] = raw["OdooProductName"].apply(TextUtils.normalize_product_clean)
                raw["ProductMappingStatus"] = "RawFromOdoo"
                raw["ProductMappingReason"] = raw["ProductMappingReason"].where(
                    raw["ProductMappingReason"].astype("string").str.strip().replace("", pd.NA).notna(),
                    "Loaded directly from Odoo sales data",
                )
                for column in ["Category", "Brand", "SubBrand", "Family", "ProductLevel", "SKU", "Size"]:
                    raw[column] = pd.NA
                raw["OdooProductID"] = pd.to_numeric(raw["OdooProductID"], errors="coerce").astype("Int64")
                raw["OdooProductTemplateID"] = pd.NA
                raw["IsActive"] = 0
                raw["ProductSource"] = self.PRODUCT_SOURCE_ODOO
                raw["IsMappedProduct"] = False
                raw = (
                    raw[self.COLUMNS]
                    .dropna(subset=["ProductKey"])
                    .drop_duplicates(subset=["ProductKey"], keep="first")
                )
                out = pd.concat([out, raw], ignore_index=True)

        return out.drop_duplicates(subset=["ProductKey"], keep="first").reset_index(drop=True)

    @classmethod
    def _odoo_catalog_rows(cls, raw_odoo_products: Optional[DataFrame]) -> DataFrame:
        if raw_odoo_products is None or raw_odoo_products.empty:
            return pd.DataFrame(columns=cls.COLUMNS)
        raw = DataFrameUtils.ensure_columns(
            raw_odoo_products.copy(),
            ["id", "product_tmpl_id_id", "product_tmpl_id", "company_id", "display_name", "name", "default_code", "active"],
        )
        out = pd.DataFrame(index=raw.index)
        out["OdooProductID"] = pd.to_numeric(raw["id"], errors="coerce").astype("Int64")
        out["OdooProductTemplateID"] = pd.to_numeric(raw.get("product_tmpl_id_id", raw.get("product_tmpl_id")), errors="coerce").astype("Int64")
        raw_name = raw["display_name"].where(
            raw["display_name"].astype("string").str.strip().replace("", pd.NA).notna(),
            raw["name"],
        )
        names = TextUtils.product_name_frame(raw_name)
        out["ProductKey"] = out["OdooProductID"].astype("string").radd("ODOO|")
        missing_key = out["OdooProductID"].isna()
        if missing_key.any():
            company = raw["company_id"].apply(TextUtils.normalize_company)
            out.loc[missing_key, "ProductKey"] = [
                ProductMapper._raw_product_key(company_value, clean_name)
                for company_value, clean_name in zip(company.loc[missing_key], names.loc[missing_key, "ProductNameClean"])
            ]
        out["Company"] = raw["company_id"].apply(TextUtils.normalize_company)
        out["ProductNameRaw"] = names["ProductNameRaw"]
        out["ProductName"] = names["ProductName"]
        out["ProductNameClean"] = names["ProductNameClean"]
        out["OdooProductName"] = names["ProductNameRaw"]
        out["OdooProductNameClean"] = names["ProductNameClean"]
        out["SKU"] = (
            raw["default_code"]
            .astype("string")
            .str.strip()
            .replace({"": pd.NA, "False": pd.NA, "false": pd.NA, "None": pd.NA, "none": pd.NA, "nan": pd.NA, "<NA>": pd.NA})
        )
        for column in ["Category", "Brand", "SubBrand", "Family", "ProductLevel", "Size"]:
            out[column] = pd.NA
        out["IsActive"] = raw["active"].fillna(False).astype(bool)
        out["ProductMappingStatus"] = "OdooProduct"
        out["ProductMappingReason"] = "Loaded directly from Odoo product.product"
        out["ProductSource"] = cls.PRODUCT_SOURCE_ODOO
        out["IsMappedProduct"] = True
        return out.dropna(subset=["ProductKey"])[cls.COLUMNS].drop_duplicates("ProductKey", keep="first").reset_index(drop=True)


class ProductCostDimensionBuilder:
    COLUMNS = [
        "OdooProductID", "OdooProductTemplateID", "Company", "ProductKey", "ProductNameRaw",
        "ProductName", "ProductNameClean", "OdooProductName", "OdooProductNameClean",
        "SKU", "ProductCost", "IsActive",
    ]

    def build(self, raw: DataFrame, dim_product: Optional[DataFrame] = None) -> DataFrame:
        """Normalize Odoo product.product cost records for model export."""
        if raw is None or raw.empty:
            return pd.DataFrame(columns=self.COLUMNS)

        raw = DataFrameUtils.ensure_columns(
            raw,
            [
                "id", "product_tmpl_id_id", "product_tmpl_id", "company_id",
                "display_name", "name", "default_code", "standard_price", "active",
            ],
        )
        out = pd.DataFrame(index=raw.index)
        out["OdooProductID"] = pd.to_numeric(raw.get("id"), errors="coerce").astype("Int64")
        out["OdooProductTemplateID"] = pd.to_numeric(
            raw.get("product_tmpl_id_id", raw.get("product_tmpl_id")), errors="coerce"
        ).astype("Int64")
        out["Company"] = raw.get("company_id", pd.Series(pd.NA, index=raw.index))
        raw_name = raw["display_name"].astype("string").fillna(raw["name"])
        names = TextUtils.product_name_frame(raw_name)
        out["ProductNameRaw"] = names["ProductNameRaw"]
        out["ProductName"] = names["ProductName"]
        out["ProductNameClean"] = names["ProductNameClean"]
        out["OdooProductName"] = names["ProductNameRaw"]
        out["OdooProductNameClean"] = names["ProductNameClean"]
        out["ProductKey"] = pd.NA
        out["SKU"] = raw.get("default_code", pd.Series(pd.NA, index=raw.index))
        out["ProductCost"] = pd.to_numeric(raw.get("standard_price"), errors="coerce")
        out["IsActive"] = raw.get("active", pd.Series(True, index=raw.index)).fillna(False).astype(bool)
        out = self._attach_official_names(out, dim_product)
        return out[self.COLUMNS].reset_index(drop=True)

    @classmethod
    def normalize_existing(cls, frame: DataFrame) -> DataFrame:
        if frame is None or frame.empty:
            return pd.DataFrame(columns=cls.COLUMNS)
        out = DataFrameUtils.ensure_columns(frame.copy(), cls.COLUMNS)
        raw_name = out["ProductNameRaw"].where(
            out["ProductNameRaw"].astype("string").str.strip().replace("", pd.NA).notna(),
            out["ProductName"],
        )
        names = TextUtils.product_name_frame(raw_name)
        out["ProductNameRaw"] = names["ProductNameRaw"]
        out["ProductName"] = names["ProductName"]
        out["ProductNameClean"] = names["ProductNameClean"]
        out["OdooProductName"] = out["OdooProductName"].where(
            out["OdooProductName"].astype("string").str.strip().replace("", pd.NA).notna(),
            out["ProductNameRaw"],
        )
        odoo_names = TextUtils.product_name_frame(out["OdooProductName"])
        out["OdooProductNameClean"] = out["OdooProductNameClean"].where(
            out["OdooProductNameClean"].astype("string").str.strip().replace("", pd.NA).notna(),
            odoo_names["ProductNameClean"],
        )
        return out[cls.COLUMNS].reset_index(drop=True)

    @staticmethod
    def _attach_official_names(out: DataFrame, dim_product: Optional[DataFrame]) -> DataFrame:
        if dim_product is None or dim_product.empty:
            return out
        lookup = DataFrameUtils.ensure_columns(
            dim_product.copy(),
            ["ProductKey", "ProductName", "ProductNameClean", "OdooProductNameClean"],
        )
        lookup["_odoo_product_name_key"] = TextUtils.clean_product_key_part(lookup["OdooProductNameClean"])
        lookup = (
            lookup[lookup["_odoo_product_name_key"].ne("")]
            .drop_duplicates(subset=["_odoo_product_name_key"], keep="first")
        )
        work = out.copy()
        work["_odoo_product_name_key"] = TextUtils.clean_product_key_part(work["OdooProductNameClean"])
        mapped = work[["_odoo_product_name_key"]].merge(
            lookup[["_odoo_product_name_key", "ProductKey", "ProductName", "ProductNameClean"]],
            on="_odoo_product_name_key",
            how="left",
            validate="m:1",
        )
        matched = mapped["ProductKey"].notna()
        work.loc[matched, "ProductKey"] = mapped.loc[matched, "ProductKey"].to_numpy()
        work.loc[matched, "ProductName"] = mapped.loc[matched, "ProductName"].to_numpy()
        work.loc[matched, "ProductNameClean"] = mapped.loc[matched, "ProductNameClean"].to_numpy()
        return work.drop(columns=["_odoo_product_name_key"], errors="ignore")


class ProductActiveFlagReconciler:
    """Refreshes PRODUCTS.xlsx[IsActive] against the live Odoo product catalog on every pipeline run.

    A product counts as active only if its ProductKey resolves to a live Odoo
    product.product record (fetched with active_test=False, so archived
    records are visible too) whose `active` flag is True. Anything that no
    longer resolves to a live Odoo record -- or resolves to an archived one --
    is flagged inactive, so Dim_Product[IsActive] (and every downstream page
    or visual that filters on it) stays correct without a manual edit.

    ProductKey values that were de-duplicated by ProductKeyUtils.ensure_unique
    (e.g. "P1|ROW|000002" for a second Odoo-name alias of the same product)
    are matched on their shared base key so every alias row gets the same flag.
    """

    _ROW_SUFFIX = re.compile(r"\|ROW\|\d{6}(\|DUP\|\d+)?$")

    @classmethod
    def _base_key(cls, series: Series) -> Series:
        return series.astype("string").fillna("").map(lambda value: cls._ROW_SUFFIX.sub("", value))

    @classmethod
    def active_base_keys(cls, dim_product_cost: DataFrame) -> set:
        if dim_product_cost is None or dim_product_cost.empty:
            return set()
        active = dim_product_cost.loc[dim_product_cost["IsActive"].fillna(False).astype(bool)]
        keys = cls._base_key(active["ProductKey"].dropna().astype(str))
        return set(keys[keys.ne("")])

    @classmethod
    def reconcile(
        cls,
        product_master: DataFrame,
        active_keys: set,
        logger: Logger,
        products_path: Optional[PathLike] = None,
    ) -> DataFrame:
        out = product_master.copy()
        if not active_keys:
            logger.warn(
                "ProductActiveFlagReconciler: no active Odoo product keys resolved this run; "
                "leaving PRODUCTS.xlsx IsActive unchanged"
            )
            return out

        base_keys = cls._base_key(out["ProductKey"])
        new_is_active = base_keys.isin(active_keys).astype(int)
        old_is_active = pd.to_numeric(out.get("IsActive"), errors="coerce").fillna(0).astype(int)
        changed = int((old_is_active != new_is_active).sum())
        out["IsActive"] = new_is_active
        if changed:
            logger.info(
                f"ProductActiveFlagReconciler: {changed} product row(s) changed IsActive vs current Odoo "
                f"catalog ({int(new_is_active.sum())} active, {int((new_is_active == 0).sum())} inactive)"
            )
        else:
            logger.info("ProductActiveFlagReconciler: IsActive already matches current Odoo catalog; no changes")

        if products_path is not None and changed:
            cls._write_back(Path(products_path), base_keys, new_is_active, logger)
        return out

    @classmethod
    def patch_dim_product(cls, dim_product: DataFrame, active_keys: set, source_filter: str) -> DataFrame:
        """Apply the same reconciliation to an already-built Dim_Product so the
        current run's export reflects it immediately, without waiting for the
        next run to pick up the rewritten PRODUCTS.xlsx."""
        if not active_keys or dim_product is None or dim_product.empty:
            return dim_product
        out = dim_product.copy()
        mask = out["ProductSource"].eq(source_filter)
        if mask.any():
            out.loc[mask, "IsActive"] = cls._base_key(out.loc[mask, "ProductKey"]).isin(active_keys).astype(int)
        return out

    @staticmethod
    def _write_back(path: Path, base_keys: Series, new_is_active: Series, logger: Logger) -> None:
        """Persist reconciled IsActive onto the PRODUCTS.xlsx source file by
        ProductKey, in place -- every other column, row, and formula (e.g. the
        SKU column) is left untouched, and no rows are dropped."""
        try:
            import openpyxl
        except ImportError:
            logger.warn("ProductActiveFlagReconciler: openpyxl not available; PRODUCTS.xlsx was not updated on disk")
            return
        if not path.exists():
            logger.warn(f"ProductActiveFlagReconciler: {path} not found; skipped write-back")
            return

        key_to_active: Dict[str, int] = dict(zip(base_keys, new_is_active))
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        try:
            key_col = header.index("ProductKey") + 1
            active_col = header.index("IsActive") + 1
        except ValueError:
            logger.warn("ProductActiveFlagReconciler: PRODUCTS.xlsx is missing ProductKey/IsActive headers; write-back skipped")
            return

        updated = 0
        for row in ws.iter_rows(min_row=2):
            key_cell = row[key_col - 1]
            if key_cell.value is None:
                continue
            new_value = key_to_active.get(str(key_cell.value))
            if new_value is not None and row[active_col - 1].value != new_value:
                row[active_col - 1].value = new_value
                updated += 1

        if updated:
            wb.save(path)
            logger.ok(f"PRODUCTS.xlsx IsActive refreshed on disk: {updated} cell(s) updated ({path})")


class ProductCostMatcher:
    @staticmethod
    def _clean(series: Series) -> Series:
        return TextUtils.clean_product_key_part(series)

    @classmethod
    def _clean_id(cls, series: Series) -> Series:
        numeric = pd.to_numeric(series, errors="coerce").astype("Int64").astype("string").fillna("")
        return numeric.mask(numeric.eq(""), cls._clean(series))

    @classmethod
    def attach(cls, sales_lines: DataFrame, dim_product_cost: DataFrame) -> DataFrame:
        """Attach cost by ID, then SKU, then cleaned product name plus company."""
        out = sales_lines.copy()
        out["product_cost"] = pd.Series(float("nan"), index=out.index, dtype="float64")
        out["ProductCostMatchKey"] = pd.NA
        out["ProductCostMissingReason"] = pd.NA
        if dim_product_cost is None or dim_product_cost.empty:
            out["ProductCostMissingReason"] = "Dim_ProductCost is empty"
            return out

        costs = DataFrameUtils.ensure_columns(
            dim_product_cost.copy(),
            ProductCostDimensionBuilder.COLUMNS,
        )
        costs["_id_key"] = cls._clean_id(costs["OdooProductID"])
        costs["_sku_key"] = cls._clean(costs["SKU"])
        costs["_name_key"] = cls._clean(costs["ProductName"])
        costs["_company_key"] = cls._clean(costs["Company"])
        costs["_row"] = 1
        costs["ProductCost"] = pd.to_numeric(costs["ProductCost"], errors="coerce")

        out["_id_key"] = (
            cls._clean_id(out["OdooProductID"])
            if "OdooProductID" in out.columns
            else pd.Series("", index=out.index, dtype="string")
        )
        if "ProductKey" in out.columns:
            out["_id_key"] = out["_id_key"].mask(
                out["_id_key"].eq(""), cls._clean(out["ProductKey"])
            )
        out["_sku_key"] = cls._clean(
            out["SKU"] if "SKU" in out.columns else pd.Series("", index=out.index)
        )
        out["_name_key"] = cls._clean(
            cls._coalesce_name(out)
        )
        company_col = next(
            (name for name in ["Company", "company_final", "company"] if name in out.columns),
            None,
        )
        out["_company_key"] = cls._clean(
            out[company_col] if company_col else pd.Series("", index=out.index)
        )

        matched = pd.Series(False, index=out.index)
        match_specs = [
            (["_id_key"], "ID"),
            (["_sku_key", "_company_key"], "SKU_COMPANY"),
            (["_sku_key"], "SKU"),
            (["_name_key", "_company_key"], "NAME_COMPANY"),
        ]
        for keys, label in match_specs:
            valid_costs = costs.copy()
            for key in keys:
                valid_costs = valid_costs[valid_costs[key].ne("")]
            lookup = (
                valid_costs.groupby(keys, dropna=False, as_index=False)
                .agg(ProductCost=("ProductCost", "mean"), _matched=("_row", "size"))
            )
            candidate = out[keys].merge(lookup, on=keys, how="left")
            can_match = (~matched) & candidate["_matched"].notna()
            out.loc[can_match, "product_cost"] = candidate.loc[can_match, "ProductCost"].to_numpy()
            out.loc[can_match, "ProductCostMatchKey"] = (
                label + ":" + out.loc[can_match, keys].astype(str).agg("|".join, axis=1)
            )
            matched |= can_match

        out.loc[~matched, "ProductCostMissingReason"] = "No matching product in Dim_ProductCost"
        out.loc[matched & pd.to_numeric(out["product_cost"], errors="coerce").isna(), "ProductCostMissingReason"] = (
            "Matched Odoo product has missing ProductCost"
        )
        return out.drop(columns=["_id_key", "_sku_key", "_name_key", "_company_key"], errors="ignore")

    @staticmethod
    def _coalesce_name(df: DataFrame) -> Series:
        result = pd.Series(pd.NA, index=df.index, dtype="string")
        for name in ["ProductName", "product_name", "ProductNameRaw", "product_name_raw", "product_name_clean"]:
            if name in df.columns:
                result = result.fillna(df[name].astype("string").str.strip().replace("", pd.NA))
        return result.apply(TextUtils.normalize_product_display)

    @classmethod
    def missing_cost_qa(
        cls,
        sales_lines: DataFrame,
        dim_product_cost: DataFrame,
        as_of_date: Optional[Any] = None,
    ) -> DataFrame:
        columns = ["Company", "ProductKey", "ProductName", "SKU", "MatchKey", "MissingReason"]
        attached = cls.attach(sales_lines, dim_product_cost)
        date_col = next(
            (name for name in ["order_date_date", "date_order", "OrderDate", "order_date"] if name in attached.columns),
            None,
        )
        if date_col:
            today = pd.Timestamp(as_of_date).normalize() if as_of_date is not None else pd.Timestamp.today().normalize()
            start = pd.Timestamp(year=today.year - 1, month=1, day=1)
            attached = attached[pd.to_datetime(attached[date_col], errors="coerce").between(start, today)]
        missing = attached[pd.to_numeric(attached["product_cost"], errors="coerce").isna()].copy()
        company_col = next((name for name in ["Company", "company_final", "company"] if name in missing.columns), None)
        missing["Company"] = missing[company_col] if company_col else pd.NA
        missing["ProductName"] = cls._coalesce_name(missing)
        missing["ProductKey"] = missing.get("ProductKey", pd.Series(pd.NA, index=missing.index))
        missing["SKU"] = missing.get("SKU", pd.Series(pd.NA, index=missing.index))
        missing["MatchKey"] = missing["ProductCostMatchKey"].fillna(
            "NAME_COMPANY:"
            + cls._clean(missing["ProductName"])
            + "|"
            + cls._clean(missing["Company"])
        )
        missing["MissingReason"] = missing["ProductCostMissingReason"]
        return (
            missing[columns]
            .drop_duplicates()
            .sort_values(["Company", "ProductName"], na_position="last")
            .reset_index(drop=True)
        )


class SegmentDimensionBuilder:
    BUSINESS_SEGMENTS = ["B2B", "B2C", "Backoffice", "Inter Company", "Unknown"]
    SEGMENT_ALIASES = {
        "B2B": "B2B",
        "B 2 B": "B2B",
        "B2C": "B2C",
        "B 2 C": "B2C",
        "BACKOFFICE": "Backoffice",
        "BACK OFFICE": "Backoffice",
        "INTER COMPANY": "Inter Company",
        "INTERCOMPANY": "Inter Company",
        "INTER-COMPANY": "Inter Company",
        "UNKNOWN": "Unknown",
        "": "Unknown",
        "NONE": "Unknown",
        "NAN": "Unknown",
        "<NA>": "Unknown",
    }

    def __init__(self, settings: PipelineSettings) -> None:
        self.settings = settings

    @classmethod
    def normalize_segment(cls, value: Any) -> str:
        if value is None or pd.isna(value):
            return "Unknown"
        text = re.sub(r"\s+", " ", str(value).strip())
        key = text.upper()
        return cls.SEGMENT_ALIASES.get(key, text if text in cls.BUSINESS_SEGMENTS else "Unknown")

    @classmethod
    def normalize_segment_series(cls, series: "pd.Series") -> "pd.Series":
        """Vectorized equivalent of normalize_segment — operates on a whole Series at once."""
        _business_set = set(cls.BUSINESS_SEGMENTS)
        normed = (
            series.astype("object")
            .fillna("Unknown")
            .astype(str)
            .str.strip()
            .str.replace(r"\s+", " ", regex=True)
        )
        upper = normed.str.upper()
        # Build extended alias map: existing aliases + pass-through for canonical business segments
        _alias: dict = dict(cls.SEGMENT_ALIASES)
        for seg in _business_set:
            _alias.setdefault(seg.upper(), seg)
        mapped = upper.map(_alias)
        return mapped.where(mapped.notna(), "Unknown")

    @classmethod
    def segment_key_map(cls) -> Dict[str, int]:
        return {segment: index + 1 for index, segment in enumerate(cls.BUSINESS_SEGMENTS)}

    def build(self, dim_salesteam: DataFrame, targets: DataFrame) -> DataFrame:
        return pd.DataFrame(
            {
                "SegmentKey": list(self.segment_key_map().values()),
                "Segment": self.BUSINESS_SEGMENTS,
            }
        )


class CustomerDimensionBuilder:
    def __init__(
        self,
        sales_org: SalesOrgMaps,
        settings: PipelineSettings,
        blocked_loader: BlockedCustomersLoader,
    ) -> None:
        self.sales_org = sales_org
        self.settings = settings
        self.blocked_loader = blocked_loader

    def build(
        self,
        df_sales: DataFrame,
        blocked_customers_path: Optional[PathLike],
        blocked_customers: Optional[DataFrame] = None,
    ) -> DataFrame:
        required = ["customer", "order_date_date", "Value"]
        missing = [col for col in required if col not in df_sales.columns]
        if missing:
            raise KeyError(f"Missing columns for Dim_Customer: {missing}")

        df = df_sales.copy()
        df = df[df["customer"].notna()].copy()

        if df.empty:
            return pd.DataFrame(columns=self._final_columns())

        df["order_date_date"] = pd.to_datetime(df["order_date_date"], errors="coerce")
        df = df[df["order_date_date"].notna()].copy()

        if df.empty:
            raise ValueError("Cannot build Dim_Customer because order_date_date has no valid dates.")

        max_date = df["order_date_date"].max()
        current_year = int(max_date.year)
        last_year = current_year - 1

        ytd_start = pd.Timestamp(year=current_year, month=1, day=1)
        lytd_start = pd.Timestamp(year=last_year, month=1, day=1)

        try:
            lytd_end = max_date.replace(year=last_year)
        except ValueError:
            lytd_end = max_date.replace(year=last_year, day=28)

        source_customer_id = df["customer_id"] if "customer_id" in df.columns else pd.Series(pd.NA, index=df.index)
        df["CustomerID"] = self._clean_customer_id(source_customer_id)
        missing_customer_id = df["CustomerID"].isna()
        if missing_customer_id.any():
            df.loc[missing_customer_id, "CustomerID"] = df.loc[missing_customer_id, "customer"].apply(self._make_synthetic_customer_id)
        grain_cols = ["CustomerID", "customer"]

        base = (
            df[grain_cols + ["order_date_date"]]
            .groupby(grain_cols, as_index=False)
            .agg(
                First_Purchase_Date=("order_date_date", "min"),
                Last_Purchase_Date=("order_date_date", "max"),
            )
        )

        company_source_col = "Company" if "Company" in df.columns else "company"
        company_map = DataFrameUtils.latest_attribute(df, grain_cols, company_source_col, "company")
        salesperson_map = DataFrameUtils.latest_attribute(df, grain_cols, "salesperson", "salesperson")

        ly_value = (
            df.loc[df["order_date_date"].dt.year == last_year, grain_cols + ["Value"]]
            .groupby(grain_cols, as_index=False)["Value"]
            .sum()
            .rename(columns={"Value": "LY_Value"})
        )

        ytd_value = (
            df.loc[
                (df["order_date_date"] >= ytd_start) & (df["order_date_date"] <= max_date),
                grain_cols + ["Value"],
            ]
            .groupby(grain_cols, as_index=False)["Value"]
            .sum()
            .rename(columns={"Value": "YTD_Value"})
        )

        full_year_values = []
        for year in [2023, 2024, 2025]:
            value_col = f"Full_{year}_Value"
            full_year_values.append(
                df.loc[df["order_date_date"].dt.year == year, grain_cols + ["Value"]]
                .groupby(grain_cols, as_index=False)["Value"]
                .sum()
                .rename(columns={"Value": value_col})
            )

        is_lytd = (
            df.loc[
                (df["order_date_date"] >= lytd_start) & (df["order_date_date"] <= lytd_end),
                grain_cols,
            ]
            .drop_duplicates()
            .assign(IsLYTD=1)
        )

        is_ytd = (
            df.loc[
                (df["order_date_date"] >= ytd_start) & (df["order_date_date"] <= max_date),
                grain_cols,
            ]
            .drop_duplicates()
            .assign(IsYTD=1)
        )

        has_history_before_lytd = (
            df.loc[df["order_date_date"] < lytd_start, grain_cols]
            .drop_duplicates()
            .assign(HasHistoryBeforeLYTD=1)
        )

        has_history_post_lytd_last_year = (
            df.loc[
                (df["order_date_date"] > lytd_end) & (df["order_date_date"] < ytd_start),
                grain_cols,
            ]
            .drop_duplicates()
            .assign(HasHistoryPostLYTDLastYear=1)
        )

        is_ly_full_year = (
            df.loc[df["order_date_date"].dt.year == last_year, grain_cols]
            .drop_duplicates()
            .assign(IsLYFullYear=1)
        )

        dim = base.merge(company_map, on=grain_cols, how="left")
        dim = dim.merge(salesperson_map, on=grain_cols, how="left")
        dim = dim.merge(ly_value, on=grain_cols, how="left")
        dim = dim.merge(ytd_value, on=grain_cols, how="left")
        for full_year_value in full_year_values:
            dim = dim.merge(full_year_value, on=grain_cols, how="left")
        dim = dim.merge(is_lytd, on=grain_cols, how="left")
        dim = dim.merge(is_ytd, on=grain_cols, how="left")
        dim = dim.merge(has_history_before_lytd, on=grain_cols, how="left")
        dim = dim.merge(has_history_post_lytd_last_year, on=grain_cols, how="left")
        dim = dim.merge(is_ly_full_year, on=grain_cols, how="left")

        dim["LY_Value"] = pd.to_numeric(dim["LY_Value"], errors="coerce").fillna(0.0)
        dim["YTD_Value"] = pd.to_numeric(dim["YTD_Value"], errors="coerce").fillna(0.0)
        for year in [2023, 2024, 2025]:
            col = f"Full_{year}_Value"
            dim[col] = pd.to_numeric(dim[col], errors="coerce").fillna(0.0)

        for col in [
            "IsLYTD",
            "IsYTD",
            "HasHistoryBeforeLYTD",
            "HasHistoryPostLYTDLastYear",
            "IsLYFullYear",
        ]:
            dim[col] = pd.to_numeric(dim[col], errors="coerce").fillna(0).astype(int)

        dim = self._add_customer_segment(df, dim, grain_cols)
        dim = self._add_sales_org_fields(dim)
        dim = self._add_customer_class(dim)
        dim["IsActiveYTD"] = (dim["IsYTD"] == 1).astype(int)

        duplicate_grain = dim.duplicated(subset=["CustomerID", "customer"]).sum()
        if duplicate_grain:
            raise ValueError(f"Dim_Customer contains duplicate CustomerID + customer grain rows: {duplicate_grain}")

        dim = self._apply_blocked_customer_data(dim, blocked_customers_path, blocked_customers)
        dim = self._add_customer_key(dim)

        dim["CustomerStatus"] = dim.apply(
            lambda row: self._calc_status(row, ytd_start, max_date),
            axis=1,
        )

        for col in self._final_columns():
            if col not in dim.columns:
                dim[col] = pd.NA

        self._validate_customer_ids(dim)
        return dim[self._final_columns()]

    @staticmethod
    def _clean_customer_id(series: Series) -> Series:
        cleaned = series.astype("string").str.strip()
        invalid = cleaned.isna() | cleaned.str.lower().isin({"", "nan", "none", "null", "<na>"})
        return cleaned.mask(invalid, pd.NA)

    @staticmethod
    def _normalize_customer_name(customer_name: Any) -> str:
        if pd.isna(customer_name):
            return "UNKNOWN"
        normalized = re.sub(r"\s+", " ", str(customer_name).strip()).upper()
        return normalized or "UNKNOWN"

    @classmethod
    def _make_synthetic_customer_id(cls, customer_name: Any) -> str:
        normalized = cls._normalize_customer_name(customer_name)
        digest = hashlib.md5(normalized.encode("utf-8")).hexdigest().upper()[:8]
        return f"CUST-{digest}"

    @staticmethod
    def _validate_customer_ids(dim: DataFrame) -> None:
        customer_id = dim["CustomerID"].astype("string").str.strip()
        invalid = customer_id.isna() | customer_id.str.lower().isin({"", "nan", "none", "null", "<na>"})
        if invalid.any():
            raise ValueError("Dim_Customer contains null or blank CustomerID values after synthetic ID generation.")

    def _add_customer_segment(
        self,
        sales: DataFrame,
        dim: DataFrame,
        grain_cols: List[str],
    ) -> DataFrame:
        if "SalesSegment" not in sales.columns:
            dim["CustomerSegment"] = "Unknown"
            return dim

        seg_base = sales[grain_cols + ["SalesSegment"]].copy()
        seg_base["SalesSegment"] = (
            seg_base["SalesSegment"]
            .astype("string")
            .str.strip()
            .str.upper()
        )

        summary = (
            seg_base.groupby(grain_cols, as_index=False)["SalesSegment"]
            .agg(lambda s: sorted(set([x for x in s.dropna().tolist() if x != ""])))
        )

        def resolve(values: Any) -> str:
            values = set(values) if isinstance(values, list) else set()
            if "B2B" in values:
                return "B2B"
            if "B2C" in values:
                return "B2C"
            return "Unknown"

        summary["CustomerSegment"] = summary["SalesSegment"].apply(resolve)
        summary = summary.drop(columns=["SalesSegment"])

        dim = dim.merge(summary, on=grain_cols, how="left")
        dim["CustomerSegment"] = dim["CustomerSegment"].fillna("Unknown")
        return dim

    def _add_sales_org_fields(self, dim: DataFrame) -> DataFrame:
        dim["salesperson"] = dim["salesperson"].astype("string")

        dim["SalesTeamKey"] = dim["salesperson"].map(self.sales_org.person_to_teamkey)
        dim["SalesTeamKey"] = (
            dim["SalesTeamKey"]
            .astype("string")
            .fillna(self.settings.unknown_team_key)
            .astype(str)
        )

        dim["DistributionChannel"] = dim["salesperson"].map(self.sales_org.person_to_channel)
        dim["DistributionChannel"] = (
            dim["DistributionChannel"]
            .astype("string")
            .fillna(self.settings.unknown_channel_name)
            .astype(str)
        )

        def branch_from_team(team_key: Any) -> str:
            if pd.isna(team_key):
                return "Unknown"
            meta = self.sales_org.teamkey_to_meta.get(str(team_key), {})
            value = meta.get("team_name")
            if pd.isna(value) or value is None or str(value).strip() == "":
                return "Unknown"
            return str(value).strip()

        dim["Branch"] = dim["SalesTeamKey"].apply(branch_from_team)
        return dim

    @staticmethod
    def _add_customer_class(dim: DataFrame) -> DataFrame:
        def classify(value: Any) -> Any:
            if pd.isna(value):
                return pd.NA
            if value > 600000:
                return "A"
            if 300000 <= value <= 599999:
                return "B"
            if 60000 <= value <= 299999:
                return "C"
            return "D"

        dim["CustomerClass_LY"] = np.where(
            dim["CustomerSegment"].isin(["B2B", "B2C"]),
            dim["LY_Value"].apply(classify),
            pd.NA,
        )
        return dim

    def _apply_blocked_customer_data(
        self,
        dim: DataFrame,
        path: Optional[PathLike],
        blocked_customers: Optional[DataFrame] = None,
    ) -> DataFrame:
        out = dim.copy()
        out["IsBlocked"] = 0
        out["BlockedDate"] = pd.NaT
        out["UnblockedDate"] = pd.NaT
        out["BlockedReason"] = pd.NA
        out["Notes"] = pd.NA

        if blocked_customers is not None:
            blocked = blocked_customers.copy()
        elif path is None or not Path(path).exists():
            return out
        else:
            blocked = self.blocked_loader.load(path)

        blocked["CustomerID"] = self._clean_customer_id(blocked["CustomerID"]) if "CustomerID" in blocked.columns else pd.NA
        matched_by_id = pd.Series(False, index=out.index)
        if "CustomerID" in blocked.columns and blocked["CustomerID"].notna().any():
            latest = (
                blocked[blocked["CustomerID"].notna()].sort_values(["CustomerID", "BlockedDate"], na_position="last")
                .drop_duplicates(subset=["CustomerID"], keep="last")
            )

            out = out.merge(
                latest[["CustomerID", "IsBlocked", "BlockedDate", "UnblockedDate", "BlockedReason", "Notes"]],
                on="CustomerID",
                how="left",
                suffixes=("", "_blk"),
            )
            matched_by_id = out["IsBlocked_blk"].notna() if "IsBlocked_blk" in out.columns else pd.Series(False, index=out.index)

        if "CustomerName" in blocked.columns and blocked["CustomerName"].notna().any():
            latest = (
                blocked.sort_values(["CustomerName", "BlockedDate"], na_position="last")
                .drop_duplicates(subset=["CustomerName"], keep="last")
            )

            name_input = out.loc[~matched_by_id].drop(columns=[f"{col}_blk" for col in ["IsBlocked", "BlockedDate", "UnblockedDate", "BlockedReason", "Notes"]], errors="ignore")
            name_match = name_input.merge(
                latest[["CustomerName", "IsBlocked", "BlockedDate", "UnblockedDate", "BlockedReason", "Notes"]],
                left_on="customer",
                right_on="CustomerName",
                how="left",
                suffixes=("", "_blk"),
            ).drop(columns=["CustomerName"], errors="ignore")
            for col in ["IsBlocked", "BlockedDate", "UnblockedDate", "BlockedReason", "Notes"]:
                blk_col = f"{col}_blk"
                if blk_col in name_match.columns:
                    out.loc[~matched_by_id, blk_col] = name_match[blk_col].to_numpy()

        for col in ["IsBlocked", "BlockedDate", "UnblockedDate", "BlockedReason", "Notes"]:
            blk_col = f"{col}_blk"
            if blk_col in out.columns:
                out[col] = out[blk_col].combine_first(out[col])
                out = out.drop(columns=[blk_col])

        out["IsBlocked"] = pd.to_numeric(out["IsBlocked"], errors="coerce").fillna(0).astype(int)
        return out

    @staticmethod
    def _add_customer_key(dim: DataFrame) -> DataFrame:
        out = dim.copy()
        out = out.sort_values(["customer", "CustomerID"], na_position="last").reset_index(drop=True)
        out.insert(0, "CustomerKey", range(1, len(out) + 1))
        return out

    @staticmethod
    def _calc_status(row: Series, ytd_start: pd.Timestamp, max_date: pd.Timestamp) -> str:
        if int(row["IsBlocked"]) == 1:
            return "Blocked"
        first_purchase = row["First_Purchase_Date"]
        is_ytd = int(row["IsYTD"])
        is_lytd = int(row["IsLYTD"])
        is_ly_full_year = int(row["IsLYFullYear"])
        has_old_history = int(row["HasHistoryBeforeLYTD"])

        if pd.notna(first_purchase) and ytd_start <= first_purchase <= max_date:
            return "New"

        # Bought in LYTD and also bought in YTD
        if is_lytd == 1 and is_ytd == 1 and is_ly_full_year == 1:
            return "Active Retained"

        # Bought before LYTD, not in LYTD, but bought in YTD
        if has_old_history == 1 and is_lytd == 0 and is_ytd == 1:
            return "Reactivated"

        # Bought in LYTD but not in YTD
        if is_lytd == 1 and is_ytd == 0:
            return "Non Active"

        return "Other"
    @staticmethod
    def _final_columns() -> List[str]:
        return [
            "CustomerKey",
            "customer",
            "CustomerID",
            "company",
            "salesperson",
            "SalesTeamKey",
            "DistributionChannel",
            "Branch",
            "First_Purchase_Date",
            "Last_Purchase_Date",
            "LY_Value",
            "YTD_Value",
            "Full_2023_Value",
            "Full_2024_Value",
            "Full_2025_Value",
            "HasHistoryBeforeLYTD",
            "HasHistoryPostLYTDLastYear",
            "IsLYTD",
            "IsYTD",
            "IsLYFullYear",
            "CustomerSegment",
            "CustomerClass_LY",
            "IsActiveYTD",
            "IsBlocked",
            "BlockedDate",
            "UnblockedDate",
            "BlockedReason",
            "Notes",
            "CustomerStatus",
        ]


class InvoiceDimensionBuilder:
    @staticmethod
    def _normalize_order_numbers(series: Series) -> Series:
        """Strip whitespace and upper-case so 'S37460 ' and 's37460' collapse to the same key."""
        return series.astype("string").str.strip().str.upper().replace("", pd.NA)

    def build(self, fact_orders: DataFrame, fact_sales: Optional[DataFrame] = None) -> DataFrame:
        if "order_number" not in fact_orders.columns:
            raise KeyError("fact_orders must contain 'order_number'")

        orders = fact_orders[["order_number"]].copy()
        orders["order_number"] = self._normalize_order_numbers(orders["order_number"])
        dim = orders.dropna(subset=["order_number"]).drop_duplicates("order_number").copy()

        if fact_sales is not None and "order_number" in fact_sales.columns:
            sales_orders = fact_sales[["order_number"]].copy()
            sales_orders["order_number"] = self._normalize_order_numbers(sales_orders["order_number"])
            extra = (
                sales_orders.dropna(subset=["order_number"]).drop_duplicates("order_number")
                .merge(dim, on="order_number", how="left", indicator=True)
            )
            extra = extra[extra["_merge"] == "left_only"][["order_number"]]
            if not extra.empty:
                dim = pd.concat([dim, extra], ignore_index=True)

        dim = dim.drop_duplicates("order_number").sort_values("order_number").reset_index(drop=True)
        dim.insert(0, "InvoiceKey", dim.index + 1)
        return dim


# ============================================================
# Facts
# ============================================================
class OffDaysFactBuilder:
    def build(self, path: PathLike, country: Optional[str]) -> DataFrame:
        df = pd.read_excel(Path(path)).copy()
        if "Date" not in df.columns:
            raise KeyError("OffDays sheet must contain a 'Date' column.")

        df["Date"] = pd.to_datetime(df["Date"], format="%d/%m/%Y", errors="coerce").dt.normalize()
        if "IsActive" in df.columns:
            df["IsActive"] = (
                df["IsActive"].astype(str).str.strip().str.lower()
                .map({"1": 1, "true": 1, "yes": 1, "y": 1, "0": 0, "false": 0, "no": 0, "n": 0})
                .fillna(0).astype(int)
            )
            df = df[df["IsActive"] == 1].copy()
        else:
            df["IsActive"] = 1

        df = DataFrameUtils.ensure_columns(df, ["OffDayType", "Country", "Company", "Branch"])
        if country and "Country" in df.columns:
            df = df[df["Country"].astype(str).str.strip().str.lower() == country.strip().lower()].copy()

        def clean_text(series: Series) -> Series:
            return series.astype("string").str.strip().replace({"": pd.NA, "nan": pd.NA, "None": pd.NA})

        df["Country"] = clean_text(df["Country"])
        df["Company"] = clean_text(df["Company"])
        df["Branch"] = clean_text(df["Branch"])
        df["OffDayType"] = clean_text(df["OffDayType"]).str.lower()

        fact = df[["Date", "OffDayType", "Country", "Company", "Branch", "IsActive"]].copy()
        fact = fact.drop_duplicates(subset=["Date", "OffDayType", "Country", "Company", "Branch"], keep="first")
        fact.insert(0, "DateKey", pd.to_datetime(fact["Date"]).dt.strftime("%Y%m%d").astype("Int64"))
        return fact.sort_values(["Date", "Country", "Company", "Branch", "OffDayType"], na_position="last").reset_index(drop=True)


class OrdersFactBuilder:
    def build(self, df_sales: DataFrame) -> DataFrame:
        if "order_number" not in df_sales.columns:
            raise KeyError("order_number column is missing before building Fact_Orders")

        aggregation = {
            "order_date_date": "min",
            "order_date": "min",
            "CustomerKey": "first",
            "customer": "first",
            "customer_id": "first",
            "salesperson": "first",
            "SalespersonKey": "first",

            "sales_team": "first",
            "sales_team_key": "first",
            "sales_team_segment": "first",
            "sales_team_city": "first",
            "sales_team_status": "first",

            "SalesTeam": "first",
            "SalesTeamKey": "first",
            "SalesSegment": "first",
            "SegmentKey": "first",
            "SalesCity": "first",
            "SalesTeamStatus": "first",

            "DistributionChannel": "first",
            "ChannelKey": "first",
            "Company": "first",
            "CompanyKey": "first",
            "line_total": "sum",
            "quantity": "sum",
        }

        for optional in ["invoice_status", "order_state"]:
            if optional in df_sales.columns:
                aggregation[optional] = "first"

        available = {k: v for k, v in aggregation.items() if k in df_sales.columns}

        # Normalize before groupby so whitespace/case variants of the same order
        # number (e.g. 'S37460 ' vs 'S37460') collapse into a single group.
        df_sales = df_sales.copy()
        df_sales["order_number"] = (
            df_sales["order_number"].astype("string").str.strip().str.upper()
        )

        fact = (
            df_sales.groupby("order_number", as_index=False)
            .agg(available)
            .rename(columns={
                "order_date_date": "OrderDate",
                "order_date": "OrderDateTime",
                "line_total": "OrderValue",
                "quantity": "OrderVolume",
                "customer_id": "CustomerID",
            })
        )

        fact["DateKey"] = pd.to_datetime(
            fact["OrderDate"], errors="coerce"
        ).dt.strftime("%Y%m%d").astype("Int64")

        return fact


class SalesLinesFactBuilder:
    COLUMNS = [
        "order_number",
        "order_date_date",
        "order_date",
        "DateKey",

        "CustomerKey",
        "customer",
        "customer_id",
        "First_Purchase_Date",
        "CustomerStatus",

        "SalespersonKey",
        "salesperson",

        "sales_team",
        "sales_team_key",
        "sales_team_segment",
        "sales_team_city",
        "sales_team_status",

        "SalesTeam",
        "SalesTeamKey",
        "SalesSegment",
        "SegmentKey",
        "SalesCity",
        "SalesTeamStatus",

        "DistributionChannel",
        "ChannelKey",

        "Company",
        "CompanyKey",

        "ProductKey",
        "OdooProductID",
        "OdooProductName",
        "OdooProductNameClean",
        "ProductJoinKey",
        "ProductLevel",
        "SKU",
        "Size",
        "product_name_raw",
        "product_name_clean",
        "product_name",
        "ProductNameRaw",
        "ProductName",
        "ProductNameClean",
        "ProductMappingStatus",
        "ProductMappingReason",
        "is_discount",

        "quantity",
        "Volume",
        "line_total",
        "Value",
        "untaxed_total",

        "invoice_status",
        "order_state",
        "InvoiceValue",
        "InvoiceClass",
        "InvoiceYear",
        "InvoiceKey",
    ]

    def build(self, df_sales: DataFrame) -> DataFrame:
        if "order_number" not in df_sales.columns:
            raise KeyError("order_number column is missing before building Fact_SalesLines")

        fact = df_sales[[col for col in self.COLUMNS if col in df_sales.columns]].copy()

        # rename for Power BI compatibility
        rename_map = {
            "Company": "company_final",
            "customer_id": "CustomerID",
            "InvoiceClass": "Invoice Class",   # <-- fix here
        }

        fact = fact.rename(columns={k: v for k, v in rename_map.items() if k in fact.columns})

        return fact

class TargetsFactBuilder:
    COLUMNS = [
        "Year", "Month", "YearMonth", "TargetDate", "DateKey",
        "TargetLevel", "Company", "CompanyKey",
        "SalesTeam", "SalesTeamKey", "Salesperson", "SalespersonKey",
        "SalesSegment", "SegmentKey", "City",
        "DistributionChannel", "ChannelKey",
        "Currency", "Target_Revenue", "Target_Volume", "ASP_LY", "ASP_ThisYear",
    ]

    def build(self, df_targets: DataFrame) -> DataFrame:
        out = DataFrameUtils.ensure_columns(df_targets, self.COLUMNS)[self.COLUMNS].copy()
        out["SalespersonKey"] = pd.to_numeric(out["SalespersonKey"], errors="coerce").fillna(0).astype(int)
        out["ChannelKey"] = pd.to_numeric(out["ChannelKey"], errors="coerce").fillna(1).astype(int)
        out["DistributionChannel"] = out["DistributionChannel"].fillna("Unknown").astype(str)
        out["SegmentKey"] = pd.to_numeric(out["SegmentKey"], errors="coerce")
        out["CompanyKey"] = pd.to_numeric(out["CompanyKey"], errors="coerce")
        return out


# ============================================================
# Product-level BCG matrix
# ============================================================
class BCGMatrixBuilder:
    GRAIN_COLUMNS = ["Company", "ProductKey", "ProductName"]
    PERIOD_METRICS = [
        "total_quantity", "total_value", "avg_unit_price", "avg_product_cost",
        "perc_gross_profit", "volume_threshold", "profit_threshold",
        "volume_class", "profit_class", "bcg_code", "bcg_class",
    ]
    COLUMNS = GRAIN_COLUMNS + ["ProductNameClean"] + [
        "total_quantity_YTD", "total_quantity_LYTD",
        "total_value_YTD", "total_value_LYTD",
        "avg_unit_price_YTD", "avg_unit_price_LYTD",
        "avg_product_cost_YTD", "avg_product_cost_LYTD",
        "perc_gross_profit_YTD", "perc_gross_profit_LYTD",
        "volume_class_YTD", "volume_class_LYTD",
        "profit_class_YTD", "profit_class_LYTD",
        "bcg_code_YTD", "bcg_code_LYTD",
        "bcg_class_YTD", "bcg_class_LYTD",
        "quantity_growth_pct", "gross_profit_change_pp", "bcg_movement",
    ]
    VALID_INVOICE_STATUSES = {"invoiced", "to invoice", "upselling"}
    VOLUME_THRESHOLDS = {"MAJAAL": 3500.0, "TIKA": 25000.0}
    PROFIT_THRESHOLD = 0.35
    BCG_RANK = {"Dogs": 1, "Cash Cows": 2, "Strategic": 3, "Stars": 4}

    @staticmethod
    def _first_available(df: DataFrame, names: Sequence[str]) -> Optional[str]:
        return next((name for name in names if name in df.columns), None)

    @staticmethod
    def _coalesce_text(df: DataFrame, names: Sequence[str]) -> Series:
        result = pd.Series(pd.NA, index=df.index, dtype="string")
        for name in names:
            if name not in df.columns:
                continue
            candidate = df[name].astype("string").str.strip().replace("", pd.NA)
            result = result.fillna(candidate)
        return result

    @classmethod
    def _filter_valid_invoices(cls, frame: DataFrame) -> DataFrame:
        if "invoice_status" not in frame.columns:
            return frame
        invoice_status = frame["invoice_status"].astype("string").str.strip().str.lower()
        return frame.loc[invoice_status.isin(cls.VALID_INVOICE_STATUSES)].copy()

    @classmethod
    def ytd_sales_value_total(cls, sales_lines: DataFrame, as_of_date: Optional[Any] = None) -> float:
        if sales_lines.empty:
            return 0.0
        date_col = cls._first_available(sales_lines, ["order_date_date", "date_order", "OrderDate", "order_date"])
        value_col = cls._first_available(sales_lines, ["Value", "line_total", "untaxed_total"])
        if date_col is None or value_col is None:
            return 0.0

        work = cls._filter_valid_invoices(sales_lines.copy())
        today = pd.Timestamp(as_of_date).normalize() if as_of_date is not None else pd.Timestamp.today().normalize()
        ytd_start = pd.Timestamp(year=today.year, month=1, day=1)
        order_date = pd.to_datetime(work[date_col], errors="coerce").dt.normalize()
        values = pd.to_numeric(work[value_col], errors="coerce")
        return float(values.loc[order_date.between(ytd_start, today)].sum())

    @classmethod
    def quality_summary(
        cls,
        sales_lines: DataFrame,
        fact_bcg: DataFrame,
        dim_segment: Optional[DataFrame] = None,
        as_of_date: Optional[Any] = None,
    ) -> Dict[str, Any]:
        duplicate_keys = ["Company", "ProductKey"]
        ytd_before = cls.ytd_sales_value_total(sales_lines, as_of_date=as_of_date)
        ytd_after = (
            float(pd.to_numeric(fact_bcg["total_value_YTD"], errors="coerce").fillna(0).sum())
            if not fact_bcg.empty and "total_value_YTD" in fact_bcg.columns
            else 0.0
        )
        segment_columns_present = [column for column in ["SegmentKey", "SalesSegment"] if column in fact_bcg.columns]
        dim_values = (
            set(dim_segment["Segment"].dropna().astype(str).str.strip())
            if dim_segment is not None and "Segment" in dim_segment.columns
            else set()
        )
        expected_segments = set(SegmentDimensionBuilder.BUSINESS_SEGMENTS)
        return {
            "row_count": int(len(fact_bcg)),
            "distinct_companies": int(fact_bcg["Company"].nunique(dropna=True)) if "Company" in fact_bcg.columns else 0,
            "duplicate_company_product_count": (
                int(fact_bcg.duplicated(duplicate_keys).sum())
                if set(duplicate_keys).issubset(fact_bcg.columns)
                else None
            ),
            "segment_columns_present": segment_columns_present,
            "segment_columns_removed": not segment_columns_present,
            "dim_segment_values": sorted(dim_values),
            "dim_segment_exact_business_values": dim_values == expected_segments if dim_segment is not None else None,
            "bcg_ytd_value": ytd_after,
            "fact_sales_valid_invoice_ytd_value": ytd_before,
            "bcg_ytd_value_delta": ytd_after - ytd_before,
            "ytd_lytd_columns_exist": {
                "total_quantity_YTD", "total_quantity_LYTD", "total_value_YTD", "total_value_LYTD",
            }.issubset(fact_bcg.columns),
        }

    @classmethod
    def build(
        cls,
        sales_lines: DataFrame,
        dim_product_cost: Optional[DataFrame] = None,
        as_of_date: Optional[Any] = None,
        dim_product: Optional[DataFrame] = None,
    ) -> DataFrame:
        """Aggregate sales lines to one wide BCG row per company and product."""
        if sales_lines.empty:
            return pd.DataFrame(columns=cls.COLUMNS)

        company_col = cls._first_available(sales_lines, ["Company", "company_final", "company"])
        date_col = cls._first_available(sales_lines, ["order_date_date", "date_order", "OrderDate", "order_date"])
        quantity_col = cls._first_available(sales_lines, ["qty_invoiced", "quantity", "Volume"])
        value_col = cls._first_available(sales_lines, ["Value", "line_total", "untaxed_total"])
        if company_col is None or date_col is None or quantity_col is None:
            raise KeyError("BCG requires company, order date, and quantity/volume columns.")

        enriched = (
            ProductCostMatcher.attach(sales_lines, dim_product_cost)
            if dim_product_cost is not None
            else sales_lines.copy()
        )
        enriched = cls._filter_valid_invoices(enriched)
        work = pd.DataFrame(index=enriched.index)
        work["Company"] = enriched[company_col].astype("string").str.strip()
        work["ProductKey"] = (
            enriched["ProductKey"].astype("string").str.strip().replace("", pd.NA)
            if "ProductKey" in enriched.columns
            else pd.Series(pd.NA, index=enriched.index, dtype="string")
        )
        fallback_raw_name = cls._coalesce_text(
            enriched, ["ProductNameRaw", "product_name_raw", "ProductName", "product_name_clean", "product_name"]
        )
        fallback_names = TextUtils.product_name_frame(fallback_raw_name)
        work["ProductName"] = fallback_names["ProductName"]
        work["ProductNameClean"] = fallback_names["ProductNameClean"]
        if dim_product is not None and not dim_product.empty and {"ProductKey", "ProductName"}.issubset(dim_product.columns):
            product_lookup = DataFrameUtils.ensure_columns(
                dim_product.copy(),
                ["ProductKey", "ProductName", "ProductNameClean"],
            )
            product_lookup["ProductKey"] = product_lookup["ProductKey"].astype("string").str.strip()
            product_lookup["ProductName"] = product_lookup["ProductName"].apply(TextUtils.normalize_product_display)
            product_lookup["ProductNameClean"] = product_lookup["ProductNameClean"].where(
                product_lookup["ProductNameClean"].astype("string").str.strip().replace("", pd.NA).notna(),
                product_lookup["ProductName"].apply(TextUtils.normalize_product_clean),
            )
            product_lookup = product_lookup.dropna(subset=["ProductKey"]).drop_duplicates(subset=["ProductKey"])
            mapped_names = work[["ProductKey"]].merge(
                product_lookup[["ProductKey", "ProductName", "ProductNameClean"]],
                on="ProductKey",
                how="left",
                validate="m:1",
            )
            work["ProductName"] = mapped_names["ProductName"].fillna(work["ProductName"]).astype("string")
            work["ProductNameClean"] = mapped_names["ProductNameClean"].fillna(work["ProductNameClean"]).astype("string")
        work["_date"] = pd.to_datetime(enriched[date_col], errors="coerce").dt.normalize()
        work["_quantity"] = pd.to_numeric(enriched[quantity_col], errors="coerce")
        work["_value"] = (
            pd.to_numeric(enriched[value_col], errors="coerce")
            if value_col
            else pd.Series(pd.NA, index=enriched.index, dtype="Float64")
        )
        work["_cost"] = (
            pd.to_numeric(enriched["product_cost"], errors="coerce")
            if "product_cost" in enriched.columns
            else pd.Series(pd.NA, index=enriched.index, dtype="Float64")
        )

        today = pd.Timestamp(as_of_date).normalize() if as_of_date is not None else pd.Timestamp.today().normalize()
        ytd_start = pd.Timestamp(year=today.year, month=1, day=1)
        lytd_start = pd.Timestamp(year=today.year - 1, month=1, day=1)
        try:
            lytd_end = today.replace(year=today.year - 1)
        except ValueError:
            lytd_end = today.replace(year=today.year - 1, day=28)

        periods = []
        for period, start, end in [("YTD", ytd_start, today), ("LYTD", lytd_start, lytd_end)]:
            selected = work.loc[work["_date"].between(start, end)].copy()
            selected["Period"] = period
            periods.append(selected)
        filtered = pd.concat(periods, ignore_index=True)
        filtered = filtered.dropna(subset=["Company", "ProductName"])
        if filtered.empty:
            return pd.DataFrame(columns=cls.COLUMNS)

        product_name_lookup = (
            filtered.dropna(subset=["ProductName"])
            .groupby(["Company", "ProductKey"], dropna=False)
            .agg(
                _CanonicalProductName=("ProductName", lambda values: values.astype("string").dropna().iloc[0] if values.astype("string").dropna().size else pd.NA),
                _CanonicalProductNameClean=("ProductNameClean", lambda values: values.astype("string").dropna().iloc[0] if values.astype("string").dropna().size else pd.NA),
            )
            .reset_index()
        )
        filtered = filtered.merge(
            product_name_lookup,
            on=["Company", "ProductKey"],
            how="left",
            validate="m:1",
        )
        filtered["ProductName"] = filtered["_CanonicalProductName"].fillna(filtered["ProductName"])
        filtered["ProductNameClean"] = filtered["_CanonicalProductNameClean"].fillna(filtered["ProductNameClean"])
        filtered = filtered.drop(columns=["_CanonicalProductName", "_CanonicalProductNameClean"])

        grain = cls.GRAIN_COLUMNS + ["Period"]
        grouped = (
            filtered.groupby(grain, dropna=False, as_index=False)
            .agg(
                ProductNameClean=("ProductNameClean", "first"),
                total_quantity=("_quantity", "sum"),
                total_value=("_value", lambda values: values.sum(min_count=1)),
                avg_product_cost=("_cost", "mean"),
            )
        )
        grouped["avg_unit_price"] = grouped["total_value"].div(
            grouped["total_quantity"].replace(0, pd.NA)
        )
        valid_profit = (
            grouped["avg_unit_price"].notna()
            & grouped["avg_unit_price"].ne(0)
            & grouped["avg_product_cost"].notna()
        )
        grouped["perc_gross_profit"] = pd.NA
        grouped.loc[valid_profit, "perc_gross_profit"] = (
            grouped.loc[valid_profit, "avg_unit_price"]
            - grouped.loc[valid_profit, "avg_product_cost"]
        ) / grouped.loc[valid_profit, "avg_unit_price"]

        company_key = grouped["Company"].astype("string").str.upper().str.strip()
        grouped["volume_threshold"] = company_key.map(cls.VOLUME_THRESHOLDS)
        grouped["profit_threshold"] = cls.PROFIT_THRESHOLD
        grouped["volume_class"] = np.where(
            grouped["total_quantity"] >= grouped["volume_threshold"], "HV", "LV"
        )
        profit = pd.to_numeric(grouped["perc_gross_profit"], errors="coerce")
        grouped["profit_class"] = np.select(
            [profit.isna(), profit.ge(cls.PROFIT_THRESHOLD)],
            ["Unknown", "HP"],
            default="LP",
        )
        grouped["bcg_code"] = grouped["volume_class"] + "/" + grouped["profit_class"]
        grouped["bcg_class"] = grouped["bcg_code"].map(
            {"HV/HP": "Stars", "HV/LP": "Cash Cows", "LV/HP": "Strategic", "LV/LP": "Dogs"}
        ).fillna("Unclassified")
        long_bcg = grouped

        # Outer-merge the periods so products sold in only one period remain
        # visible and can be classified as New or Lost.
        period_frames = []
        for period in ["YTD", "LYTD"]:
            period_frame = long_bcg.loc[
                long_bcg["Period"].eq(period),
                cls.GRAIN_COLUMNS + ["ProductNameClean"] + cls.PERIOD_METRICS,
            ].copy()
            period_frame = period_frame.rename(
                columns={metric: f"{metric}_{period}" for metric in cls.PERIOD_METRICS}
            )
            period_frames.append(period_frame)
        wide = period_frames[0].merge(
            period_frames[1],
            on=cls.GRAIN_COLUMNS,
            how="outer",
            validate="1:1",
        )
        wide["ProductNameClean"] = wide["ProductNameClean_x"].fillna(wide["ProductNameClean_y"])
        wide = wide.drop(columns=["ProductNameClean_x", "ProductNameClean_y"], errors="ignore")

        lytd_quantity = pd.to_numeric(wide["total_quantity_LYTD"], errors="coerce")
        ytd_quantity = pd.to_numeric(wide["total_quantity_YTD"], errors="coerce")
        wide["quantity_growth_pct"] = (ytd_quantity - lytd_quantity).div(
            lytd_quantity.where(lytd_quantity.ne(0))
        )
        wide["gross_profit_change_pp"] = (
            pd.to_numeric(wide["perc_gross_profit_YTD"], errors="coerce")
            - pd.to_numeric(wide["perc_gross_profit_LYTD"], errors="coerce")
        )

        ytd_class = wide["bcg_class_YTD"].astype("string")
        lytd_class = wide["bcg_class_LYTD"].astype("string")
        ytd_rank = ytd_class.map(cls.BCG_RANK)
        lytd_rank = lytd_class.map(cls.BCG_RANK)
        wide["bcg_movement"] = pd.Series(pd.NA, index=wide.index, dtype="string")
        wide.loc[ytd_class.notna() & lytd_class.notna() & ytd_class.eq(lytd_class), "bcg_movement"] = "Stable"
        wide.loc[ytd_class.notna() & lytd_class.isna(), "bcg_movement"] = "New"
        wide.loc[ytd_class.isna() & lytd_class.notna(), "bcg_movement"] = "Lost"
        wide.loc[ytd_rank.gt(lytd_rank), "bcg_movement"] = "Improved"
        wide.loc[ytd_rank.lt(lytd_rank), "bcg_movement"] = "Declined"

        return (
            wide[cls.COLUMNS]
            .sort_values(["Company", "bcg_class_YTD", "ProductName"], na_position="last")
            .reset_index(drop=True)
        )


# ============================================================
# QA + exporter
# ============================================================
class QAService:
    UNMAPPED_PRODUCT_COLUMNS = [
        "Company", "OdooProductName", "OdooProductNameClean", "ProductKey",
        "Source", "FirstSeenDate", "RefreshDate", "ProductMappingStatus", "Reason",
    ]
    PRODUCT_MAPPING_CHECK_COLUMNS = ["CheckName", "MetricValue", "Status", "Notes"]

    @staticmethod
    def _clean_key_part(series: Series) -> Series:
        """Uppercase, trim, collapse spaces, and represent null as blank."""
        return TextUtils.clean_product_key_part(series)

    @classmethod
    def unmapped_products(cls, df_sales: DataFrame, product_master: DataFrame) -> DataFrame:
        if df_sales is None or df_sales.empty:
            return pd.DataFrame(columns=cls.UNMAPPED_PRODUCT_COLUMNS)
        out = DataFrameUtils.ensure_columns(
            df_sales.copy(),
            cls.UNMAPPED_PRODUCT_COLUMNS
            + ["ProductMappingReason", "ProductName", "company_final", "company", "order_date_date"],
        )
        out = out[out["ProductMappingStatus"].astype("string").eq("RawFromOdoo")].copy()
        if out.empty:
            return pd.DataFrame(columns=cls.UNMAPPED_PRODUCT_COLUMNS)
        if "Company" not in out.columns and "company_final" in out.columns:
            out["Company"] = out["company_final"]
        out["Company"] = out["Company"].where(
            out["Company"].astype("string").str.strip().replace("", pd.NA).notna(),
            out["company_final"],
        )
        out["Company"] = out["Company"].where(
            out["Company"].astype("string").str.strip().replace("", pd.NA).notna(),
            out["company"],
        )
        out["OdooProductName"] = out["OdooProductName"].where(
            out["OdooProductName"].astype("string").str.strip().replace("", pd.NA).notna(),
            out["ProductName"],
        )
        out["OdooProductNameClean"] = out["OdooProductNameClean"].where(
            out["OdooProductNameClean"].astype("string").str.strip().replace("", pd.NA).notna(),
            out["OdooProductName"].apply(TextUtils.normalize_product_clean),
        )
        out["Source"] = out["Source"].fillna("Odoo Sales")
        out["FirstSeenDate"] = pd.to_datetime(out["FirstSeenDate"], errors="coerce")
        order_dates = pd.to_datetime(out["order_date_date"], errors="coerce")
        out["FirstSeenDate"] = out["FirstSeenDate"].where(out["FirstSeenDate"].notna(), order_dates)
        out["RefreshDate"] = pd.to_datetime(out["RefreshDate"], errors="coerce")
        refresh_date = pd.Timestamp.today().normalize()
        out["RefreshDate"] = out["RefreshDate"].where(out["RefreshDate"].notna(), refresh_date)
        out["Reason"] = out["Reason"].where(
            out["Reason"].astype("string").str.strip().replace("", pd.NA).notna(),
            out["ProductMappingReason"],
        )
        grouped = (
            out[cls.UNMAPPED_PRODUCT_COLUMNS]
            .sort_values(["Company", "OdooProductNameClean", "FirstSeenDate"], na_position="last")
            .groupby(["Company", "OdooProductName", "OdooProductNameClean", "ProductKey", "Source"], dropna=False, as_index=False)
            .agg(
                FirstSeenDate=("FirstSeenDate", "min"),
                RefreshDate=("RefreshDate", "max"),
                ProductMappingStatus=("ProductMappingStatus", "first"),
                Reason=("Reason", "first"),
            )
        )
        return grouped[cls.UNMAPPED_PRODUCT_COLUMNS].reset_index(drop=True)

    @classmethod
    def product_mapping_checks(
        cls,
        dim_product: DataFrame,
        product_master: DataFrame,
        qa_unmapped: Optional[DataFrame] = None,
    ) -> DataFrame:
        dim = DataFrameUtils.ensure_columns(
            dim_product.copy(),
            [
                "ProductKey", "Category", "Brand", "SubBrand", "Family", "ProductLevel",
                "SKU", "Size", "ProductSource", "IsMappedProduct",
            ],
        )
        master = DataFrameUtils.ensure_columns(product_master.copy(), ["ProductName", "ProductKey"])
        master = master[master["ProductName"].astype("string").str.strip().replace("", pd.NA).notna()].reset_index(drop=True)
        master = ProductKeyUtils.ensure_unique(master)
        mapped = dim["IsMappedProduct"].astype("string").str.lower().isin(["true", "1"])
        input_source = dim["ProductSource"].astype("string").eq(ProductDimensionBuilder.PRODUCT_SOURCE_INPUT)
        placeholders = dim["ProductSource"].astype("string").eq(ProductDimensionBuilder.PRODUCT_SOURCE_UNMAPPED) | ~mapped
        enriched = mapped | input_source
        master_keys = set(master["ProductKey"].dropna().astype("string"))
        enriched_keys = set(dim.loc[enriched, "ProductKey"].dropna().astype("string"))
        descriptive_columns = ["Category", "Brand", "SubBrand", "Family", "ProductLevel", "SKU", "Size"]
        placeholder_descriptive = dim.loc[placeholders, descriptive_columns]
        filled_placeholder_attrs = 0
        if not placeholder_descriptive.empty:
            filled_placeholder_attrs = int(
                placeholder_descriptive.apply(
                    lambda col: col.astype("string").str.strip().replace("", pd.NA).notna()
                ).sum().sum()
            )
        missing_from_master = sorted(enriched_keys - master_keys)
        rows = [
            ("Mapped products", int(enriched.sum()), "INFO", ""),
            (
                "Unmapped Odoo products",
                0 if qa_unmapped is None else len(qa_unmapped),
                "INFO",
                "",
            ),
            ("Dim_Product rows from input master", int(input_source.sum()), "INFO", ""),
            ("Placeholder unmapped rows", int(placeholders.sum()), "INFO", ""),
            (
                "Unmapped placeholders have no enriched attributes",
                filled_placeholder_attrs,
                "PASS" if filled_placeholder_attrs == 0 else "FAIL",
                "Counts filled Category/Brand/SubBrand/Family/ProductLevel/SKU/Size cells on placeholder rows",
            ),
            (
                "Every enriched Dim_Product row exists in PRODUCTS.xlsx",
                len(missing_from_master),
                "PASS" if not missing_from_master else "FAIL",
                ", ".join(missing_from_master[:10]),
            ),
        ]
        return pd.DataFrame(rows, columns=cls.PRODUCT_MAPPING_CHECK_COLUMNS)


class WorkbookExporter:
    EXCEL_EPOCH = pd.Timestamp("1899-12-30")

    def __init__(self, logger: Logger) -> None:
        self.logger = logger

    @classmethod
    def _excel_datetime_serial(cls, value: Any) -> float:
        timestamp = pd.Timestamp(value).tz_localize(None)
        return float((timestamp - cls.EXCEL_EPOCH) / pd.Timedelta(days=1))

    @classmethod
    def _excel_cell_value(cls, value: Any) -> Any:
        if pd.isna(value):
            return None
        if isinstance(value, pd.Timestamp):
            return cls._excel_datetime_serial(value)
        if isinstance(value, dt.datetime):
            return cls._excel_datetime_serial(value)
        if isinstance(value, dt.date):
            return cls._excel_datetime_serial(pd.Timestamp(value))
        if hasattr(value, "item"):
            try:
                item = value.item()
            except ValueError:
                return value
            if isinstance(item, (dt.datetime, dt.date)):
                return cls._excel_datetime_serial(item)
            return item
        return value

    @classmethod
    def _excel_rows(cls, rows: list[list[Any]]) -> tuple[tuple[Any, ...], ...]:
        return tuple(tuple(cls._excel_cell_value(value) for value in row) for row in rows)

    @staticmethod
    def _is_datetime_series(series: Series) -> bool:
        if pd.api.types.is_datetime64_any_dtype(series):
            return True
        sample = series.dropna().head(25)
        return any(isinstance(value, (pd.Timestamp, dt.datetime, dt.date)) for value in sample.tolist())

    @staticmethod
    def _is_date_only_column(column_name: str, series: Series) -> bool:
        lowered = column_name.lower()
        if lowered.endswith("_date_date") or lowered in {"date", "targetdate", "orderdate", "salesorderdate", "quotationdate"}:
            return True
        values = pd.to_datetime(series.dropna().head(100), errors="coerce").dropna()
        return bool(not values.empty and values.dt.time.eq(dt.time()).all())

    @staticmethod
    def _get_or_create_excel_sheet(workbook: Any, sheet_name: str) -> Any:
        try:
            return workbook.Worksheets(sheet_name)
        except Exception:
            sheet = workbook.Worksheets.Add(After=workbook.Worksheets(workbook.Worksheets.Count))
            sheet.Name = sheet_name
            return sheet

    @classmethod
    def _write_frame_to_excel_sheet(cls, sheet: Any, frame: DataFrame) -> None:
        for index in range(sheet.ListObjects.Count, 0, -1):
            sheet.ListObjects(index).Delete()
        sheet.UsedRange.Clear()
        sheet.Cells.Clear()

        column_count = len(frame.columns)
        if column_count == 0:
            return

        header = cls._excel_rows([frame.columns.tolist()])
        sheet.Range(sheet.Cells(1, 1), sheet.Cells(1, column_count)).Value = header

        if frame.empty:
            return

        cleaned = frame.astype(object).where(pd.notna(frame), None)
        chunk_size = 5000
        for start in range(0, len(cleaned), chunk_size):
            chunk = cleaned.iloc[start:start + chunk_size].values.tolist()
            values = cls._excel_rows(chunk)
            first_row = start + 2
            last_row = first_row + len(values) - 1
            sheet.Range(sheet.Cells(first_row, 1), sheet.Cells(last_row, column_count)).Value = values

        for column_number, column_name in enumerate(frame.columns, start=1):
            series = frame[column_name]
            if not cls._is_datetime_series(series):
                continue
            number_format = "yyyy-mm-dd" if cls._is_date_only_column(str(column_name), series) else "yyyy-mm-dd hh:mm:ss"
            sheet.Range(sheet.Cells(2, column_number), sheet.Cells(len(frame) + 1, column_number)).NumberFormat = number_format

    def _export_to_open_excel(self, sheets: Dict[str, DataFrame], path: Path) -> Path:
        import win32com.client

        excel = win32com.client.GetActiveObject("Excel.Application")
        workbook = next(
            (
                book
                for book in excel.Workbooks
                if Path(book.FullName).resolve() == path.resolve()
            ),
            None,
        )
        if workbook is None:
            raise RuntimeError(f"Workbook is locked and no open Excel instance was found for: {path}")

        for sheet_name, frame in sheets.items():
            if sheet_name in {"QA_MissingProductCost"}:
                continue
            sheet = self._get_or_create_excel_sheet(workbook, sheet_name)
            self._write_frame_to_excel_sheet(sheet, frame)
        workbook.Save()
        self.logger.ok(f"Exported workbook through open Excel instance: {path}")
        return path

    @staticmethod
    def _alternate_workbook_path(path: Path) -> Path:
        timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        candidate = path.with_name(f"{path.stem}_{timestamp}{path.suffix}")
        suffix = 2
        while candidate.exists():
            candidate = path.with_name(f"{path.stem}_{timestamp}_{suffix}{path.suffix}")
            suffix += 1
        return candidate

    def _write_workbook_file(
        self,
        path: Path,
        sheets: Dict[str, DataFrame],
        existing_sheet_order: list[str],
        preserved_sheets: Dict[str, DataFrame],
        separate_qa_sheets: set[str],
    ) -> None:
        with pd.ExcelWriter(path, engine="xlsxwriter") as writer:
            written: set[str] = set()
            for sheet_name in existing_sheet_order:
                if sheet_name in separate_qa_sheets:
                    continue
                df = sheets.get(sheet_name, preserved_sheets.get(sheet_name))
                if df is None:
                    continue
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                written.add(sheet_name)
            for sheet_name, df in sheets.items():
                if sheet_name in separate_qa_sheets or sheet_name in written:
                    continue
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    def export(self, sheets: Dict[str, DataFrame], path: PathLike) -> Path:
        out = Path(path)
        out.parent.mkdir(parents=True, exist_ok=True)
        separate_qa_sheets = {"QA_MissingProductCost"}
        existing_sheet_order: list[str] = []
        preserved_sheets: Dict[str, DataFrame] = {}
        if out.exists():
            try:
                existing_book = pd.ExcelFile(out)
                existing_sheet_order = existing_book.sheet_names
                for sheet_name in existing_sheet_order:
                    if sheet_name not in sheets and sheet_name not in separate_qa_sheets:
                        preserved_sheets[sheet_name] = pd.read_excel(existing_book, sheet_name=sheet_name)
            except Exception as exc:  # noqa: BLE001
                self.logger.warn(f"Could not read existing workbook sheets for preservation: {exc}")
        qa_missing_cost = sheets.get(
            "QA_MissingProductCost",
            pd.DataFrame(columns=["Company", "ProductKey", "ProductName", "SKU", "MatchKey", "MissingReason"]),
        )
        qa_missing_cost.to_excel(out.parent / "QA_MissingProductCost.xlsx", index=False)
        (out.parent / "QA_AmbiguousProductMapping.xlsx").unlink(missing_ok=True)
        (out.parent / "QA_UnmappedProducts.xlsx").unlink(missing_ok=True)

        try:
            self._write_workbook_file(out, sheets, existing_sheet_order, preserved_sheets, separate_qa_sheets)
        except PermissionError:
            self.logger.warn(f"Workbook is open or locked; updating through Excel instead: {out}")
            try:
                return self._export_to_open_excel(sheets, out)
            except Exception as exc:  # noqa: BLE001
                alternate = self._alternate_workbook_path(out)
                self.logger.warn(
                    f"Open Excel save failed ({exc}); exporting alternate workbook instead: {alternate}"
                )
                self._write_workbook_file(alternate, sheets, existing_sheet_order, preserved_sheets, separate_qa_sheets)
                self.logger.ok(f"Exported alternate workbook: {alternate}")
                return alternate
        self.logger.ok(f"Exported workbook: {out}")
        return out


# ============================================================
# Pipeline
# ============================================================
class SalesModelPipeline:
    def __init__(self, config: PipelineConfig) -> None:
        self.config = config
        self.logger = Logger(enabled=config.settings.verbose)

        self.sales_loader = SalesLoader(self.logger)
        self.targets_loader = TargetsLoader(self.logger)
        self.blocked_loader = BlockedCustomersLoader(self.logger)
        self.sales_org_repo = SalesOrgRepository(self.logger)
        self.product_loader = ProductMasterLoader(
            logger=self.logger,
            export_conflicts=config.settings.export_product_conflicts,
            base_dir=config.paths.base_dir,
        )
        self.sales_cleaner = SalesCleaner()
        self.exporter = WorkbookExporter(self.logger)

    def run(self) -> Path:
        paths = self.config.paths
        settings = self.config.settings

        if not paths.blocked_customers_path.exists():
            self.logger.info(f"Blocked customers file not found: {paths.blocked_customers_path}")
            self.blocked_loader.export_template(paths.blocked_customers_path)

        sales_org = self.sales_org_repo.load(paths.master_path)
        sales_org_enricher = SalesOrgEnricher(sales_org)

        sales = self.sales_loader.load(paths.sales_path)
        targets = self.targets_loader.load(paths.targets_path)
        product_master = self.product_loader.load(paths.products_path)

        sales = self.sales_cleaner.clean_order_numbers(sales)
        sales = self.sales_cleaner.clean_customer(sales)
        sales = self.sales_cleaner.clean_product_names(sales)
        sales = ProductMapper.attach(sales, product_master)
        sales = self.sales_cleaner.clean_quantity(sales)
        sales = self.sales_cleaner.clean_value(sales)
        sales = self.sales_cleaner.clean_order_dates(sales)
        sales = self.sales_cleaner.filter_dashboard_rows(sales)
        sales = self.sales_cleaner.add_first_purchase_date(sales)
        sales = self.sales_cleaner.clean_salesperson(sales)
        sales = sales_org_enricher.enrich_sales(sales)
        sales = self.sales_cleaner.clean_company(sales)
        sales = self.sales_cleaner.add_invoice_summary(sales)
        sales["DateKey"] = pd.to_datetime(
            sales["order_date_date"], errors="coerce"
        ).dt.strftime("%Y%m%d").astype("Int64")
        targets = sales_org_enricher.normalize_targets(targets)
        targets["TargetDate"] = pd.to_datetime(targets["TargetDate"], errors="coerce")
        targets["DateKey"] = targets["TargetDate"].dt.strftime("%Y%m%d").astype("Int64")

        dim_product = ProductDimensionBuilder().build(product_master, sales)
        dim_salesperson = SalespersonDimensionBuilder(sales_org, settings).build(sales)
        dim_channel = DistributionChannelDimensionBuilder().build(sales_org.people_active, targets)
        dim_salesteam = SalesTeamDimensionBuilder(sales_org, settings).build()
        dim_date = DateDimensionBuilder(settings.weekly_rest_day_name).build(sales)
        dim_company = CompanyDimensionBuilder().build(sales)
        dim_segment = SegmentDimensionBuilder(settings).build(dim_salesteam, targets)

        sales = self._attach_salesperson_fields(sales, dim_salesperson)

        sales = DataFrameUtils.add_key_from_dimension(
            sales,
            "DistributionChannel",
            dim_channel,
            "DistributionChannel",
            "ChannelKey",
        )
        sales = DataFrameUtils.add_key_from_dimension(
            sales,
            "Company",
            dim_company,
            "Company",
            "CompanyKey",
        )

        targets = DataFrameUtils.add_key_from_dimension(
            targets,
            "DistributionChannel",
            dim_channel,
            "DistributionChannel",
            "ChannelKey",
        )
        targets = DataFrameUtils.add_key_from_dimension(
            targets,
            "Company",
            dim_company,
            "Company",
            "CompanyKey",
        )
        targets = DataFrameUtils.add_key_from_dimension(
            targets,
            "SalesSegment",
            dim_segment,
            "Segment",
            "SegmentKey",
        )

        salesperson_key_map = dict(
            zip(
                dim_salesperson["salesperson"].astype(str),
                dim_salesperson["SalespersonKey"].astype(int),
            )
        )
        targets["SalespersonKey"] = targets["Salesperson"].astype(str).map(salesperson_key_map)
        targets["SalespersonKey"] = pd.to_numeric(
            targets["SalespersonKey"], errors="coerce"
        ).fillna(0).astype(int)

        dim_customer = CustomerDimensionBuilder(
            sales_org,
            settings,
            self.blocked_loader,
        ).build(
            sales,
            blocked_customers_path=paths.blocked_customers_path,
        )

        sales = self._attach_customer_keys(sales, dim_customer)

        fact_offdays = OffDaysFactBuilder().build(
            paths.offdays_path,
            settings.offdays_country,
        )
        fact_orders = OrdersFactBuilder().build(sales)

        dim_invoice = InvoiceDimensionBuilder().build(fact_orders)

        sales = DataFrameUtils.add_key_from_dimension(
            sales,
            "order_number",
            dim_invoice,
            "order_number",
            "InvoiceKey",
        )
        fact_orders = DataFrameUtils.add_key_from_dimension(
            fact_orders,
            "order_number",
            dim_invoice,
            "order_number",
            "InvoiceKey",
        )

        fact_sales = SalesLinesFactBuilder().build(sales)
        fact_targets = TargetsFactBuilder().build(targets)
        if "order_number" not in fact_sales.columns:
            raise KeyError("Fact_SalesLines is missing required column: order_number")
        if "sales_team" not in fact_sales.columns:
            raise KeyError("Fact_SalesLines is missing required column: sales_team")
        if "order_number" not in fact_orders.columns:
            raise KeyError("Fact_Orders is missing required column: order_number")
        if "sales_team" not in fact_orders.columns:
            raise KeyError("Fact_Orders is missing required column: sales_team")

        workbook = {
            "Fact_SalesLines": fact_sales,
            "Fact_Orders": fact_orders,
            "Fact_Targets": fact_targets,
            "Fact_OffDays": fact_offdays,
            "Dim_Date": dim_date,
            "Dim_Customer": dim_customer,
            "Dim_Salesperson": dim_salesperson,
            "Dim_SalesTeam": dim_salesteam,
            "Dim_Company": dim_company,
            "Dim_Product": dim_product,
            "Dim_DistributionChannel": dim_channel,
            "Dim_Segment": dim_segment,
            "Dim_Invoice": dim_invoice,
        }

        return self.exporter.export(workbook, paths.output_path)

    @staticmethod
    def _attach_salesperson_fields(df_sales: DataFrame, dim_salesperson: DataFrame) -> DataFrame:
        out = df_sales.copy()

        lookup_cols = [
            "salesperson",
            "SalespersonKey",
            "DistributionChannel",
            "SalesTeamKey",
        ]
        lookup = (
            dim_salesperson[lookup_cols]
            .dropna(subset=["salesperson"])
            .drop_duplicates(subset=["salesperson"])
        )

        for col in ["SalespersonKey", "DistributionChannel"]:
            if col in out.columns:
                out = out.drop(columns=[col])

        out = out.merge(
            lookup,
            on="salesperson",
            how="left",
            validate="m:1",
            suffixes=("", "_dim"),
        )

        if "SalesTeamKey_dim" in out.columns:
            if "SalesTeamKey" in out.columns:
                out["SalesTeamKey"] = out["SalesTeamKey"].combine_first(out["SalesTeamKey_dim"])
            else:
                out["SalesTeamKey"] = out["SalesTeamKey_dim"]
            out = out.drop(columns=["SalesTeamKey_dim"])

        out["SalespersonKey"] = pd.to_numeric(
            out["SalespersonKey"], errors="coerce"
        ).fillna(0).astype(int)

        out["DistributionChannel"] = (
            out["DistributionChannel"]
            .fillna("Unknown")
            .astype(str)
            .str.strip()
        )
        out.loc[out["DistributionChannel"] == "", "DistributionChannel"] = "Unknown"

        return out

    @staticmethod
    def _attach_customer_keys(df_sales: DataFrame, dim_customer: DataFrame) -> DataFrame:
        out = df_sales.copy()

        for col in ["CustomerKey", "CustomerStatus"]:
            if col in out.columns:
                out = out.drop(columns=[col])

        source_customer_id = out["customer_id"] if "customer_id" in out.columns else pd.Series(pd.NA, index=out.index)
        out["_CustomerIDJoin"] = CustomerDimensionBuilder._clean_customer_id(source_customer_id)
        missing_customer_id = out["_CustomerIDJoin"].isna()
        if missing_customer_id.any():
            out.loc[missing_customer_id, "_CustomerIDJoin"] = out.loc[missing_customer_id, "customer"].apply(CustomerDimensionBuilder._make_synthetic_customer_id)
        lookup = (
            dim_customer[["CustomerID", "CustomerKey", "CustomerStatus"]]
            .dropna(subset=["CustomerID"])
            .drop_duplicates(subset=["CustomerID"])
        )
        out = out.merge(
            lookup,
            left_on="_CustomerIDJoin",
            right_on="CustomerID",
            how="left",
            validate="m:1",
        ).drop(columns=["CustomerID", "_CustomerIDJoin"], errors="ignore")

        out["CustomerKey"] = pd.to_numeric(out["CustomerKey"], errors="coerce").astype("Int64")
        return out


# ============================================================
# Public entrypoint
# ============================================================
def run_pipeline(
    sales_path: Optional[PathLike] = None,
    targets_path: Optional[PathLike] = None,
    master_path: Optional[PathLike] = None,
    output_path: Optional[PathLike] = None,
    offdays_path: Optional[PathLike] = None,
    products_path: Optional[PathLike] = None,
    blocked_customers_path: Optional[PathLike] = None,
    base_dir: PathLike = r"C:\Users\Lenovo\Desktop\Sales Data",
    verbose: bool = True,
) -> Path:
    paths = PipelinePaths.from_base_dir(base_dir)
    if sales_path is not None:
        paths = replace(paths, sales_path=Path(sales_path))
    if targets_path is not None:
        paths = replace(paths, targets_path=Path(targets_path))
    if master_path is not None:
        paths = replace(paths, master_path=Path(master_path))
    if output_path is not None:
        paths = replace(paths, output_path=Path(output_path))
    if offdays_path is not None:
        paths = replace(paths, offdays_path=Path(offdays_path))
    if products_path is not None:
        paths = replace(paths, products_path=Path(products_path))
    if blocked_customers_path is not None:
        paths = replace(paths, blocked_customers_path=Path(blocked_customers_path))

    settings = PipelineSettings(verbose=verbose)
    config = PipelineConfig(paths=paths, settings=settings)
    return SalesModelPipeline(config).run()


if __name__ == "__main__":
    run_pipeline()
