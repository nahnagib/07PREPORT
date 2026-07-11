from pathlib import Path
from types import SimpleNamespace

import pandas as pd

from config.settings import Settings
from sales_pipeline.legacy_transform import (
    BCGMatrixBuilder,
    Logger,
    PipelineSettings,
    ProductActiveFlagReconciler,
    ProductCostDimensionBuilder,
    ProductCostMatcher,
    ProductDimensionBuilder,
    ProductMapper,
    QAService,
    SegmentDimensionBuilder,
    SalesCleaner,
    SalesLinesFactBuilder,
    WorkbookExporter,
    normalize_product_name,
)
from sales_pipeline.pipeline import PowerBISalesPipeline
from sales_pipeline.validation import ModelValidator


EXPECTED_PRODUCT_COLUMNS = [
    "ProductKey",
    "Company",
    "Category",
    "Brand",
    "SubBrand",
    "Family",
    "ProductNameRaw",
    "ProductName",
    "ProductNameClean",
    "OdooProductName",
    "OdooProductNameClean",
    "ProductLevel",
    "SKU",
    "Size",
    "IsActive",
    "ProductMappingStatus",
    "ProductMappingReason",
    "ProductSource",
    "IsMappedProduct",
]


def test_dim_product_preserves_all_manual_master_rows_and_null_hierarchy() -> None:
    product_master = pd.DataFrame(
        [
            {
                "ProductKey": "P1",
                "Company": "A",
                "Category": pd.NA,
                "Brand": pd.NA,
                "SubBrand": pd.NA,
                "Family": pd.NA,
                    "ProductName": "Official Product One",
                    "OdooProductName": "Product One",
                "ProductLevel": "SKU",
                "SKU": "SKU1",
                "Size": pd.NA,
                "IsActive": 1,
            },
            {
                "ProductKey": "P1",
                "Company": "A",
                    "ProductName": "Official Product One Alias",
                    "OdooProductName": "Product One Alias",
                "ProductLevel": "SKU",
                "SKU": "SKU1",
                "IsActive": 1,
            },
        ]
    )

    result = ProductDimensionBuilder().build(product_master)

    assert len(result) == len(product_master)
    assert result.columns.tolist() == EXPECTED_PRODUCT_COLUMNS
    assert result["ProductKey"].tolist() == ["P1", "P1|ROW|000002"]
    assert result["ProductKey"].is_unique
    assert pd.isna(result.loc[0, "Category"])
    assert result["ProductName"].tolist() == ["Official Product One", "Official Product One Alias"]
    assert result["OdooProductNameClean"].tolist() == ["PRODUCT ONE", "PRODUCT ONE ALIAS"]
    assert result["ProductMappingReason"].tolist() == ["Loaded from PRODUCTS.xlsx", "Loaded from PRODUCTS.xlsx"]
    assert result["ProductSource"].tolist() == ["Input Master", "Input Master"]
    assert result["IsMappedProduct"].tolist() == [True, True]


def test_product_mapping_qa_reports_only_unmapped_odoo_products() -> None:
    sales = pd.DataFrame(
        [
            {"OdooProductName": "Odoo One", "OdooProductNameClean": "ODOO ONE", "ProductName": "Official One", "ProductKey": "P1", "Company": "Majaal", "ProductMappingStatus": "MatchedByOdooProductName", "ProductMappingReason": "matched", "order_date_date": "2026-01-05"},
            {"OdooProductName": "Raw Item", "OdooProductNameClean": "RAW ITEM", "ProductName": "Raw Item", "ProductKey": "RAW_FROM_ODOO|ABC", "Company": "Majaal", "ProductMappingStatus": "RawFromOdoo", "ProductMappingReason": "raw", "order_date_date": "2026-01-03"},
        ]
    )

    result = QAService.unmapped_products(sales, pd.DataFrame())

    assert result["OdooProductName"].tolist() == ["Raw Item"]
    assert result["ProductKey"].tolist() == ["RAW_FROM_ODOO|ABC"]
    assert result["Source"].tolist() == ["Odoo Sales"]
    assert result["FirstSeenDate"].tolist() == [pd.Timestamp("2026-01-03")]
    assert result["ProductMappingStatus"].tolist() == ["RawFromOdoo"]
    assert result["Reason"].tolist() == ["raw"]


def test_product_mapper_matches_only_odoo_name_and_keeps_unmapped_sparse() -> None:
    sales = pd.DataFrame(
        [
            {"product_name_raw": "Odoo Product One", "product_name_clean": "ODOO PRODUCT ONE"},
            {"product_name_raw": "Unmapped Odoo Product", "product_name_clean": "UNMAPPED ODOO PRODUCT"},
        ]
    )
    product_master = pd.DataFrame(
        [
            {
                "ProductKey": "P1",
                "ProductName": "Official Product One",
                "OdooProductName": "Odoo Product One",
                "SKU": "SKU1",
                "ProductLevel": "SKU",
                "Company": "A",
                "IsActive": 1,
            }
        ]
    )

    result = ProductMapper.attach(sales, product_master)

    assert result["ProductKey"].tolist()[0] == "P1"
    assert result["ProductName"].tolist()[0] == "Official Product One"
    assert result["product_name"].tolist()[0] == "Official Product One"
    assert result["ProductMappingStatus"].tolist()[0] == "MatchedByOdooProductName"
    assert result["ProductKey"].tolist()[1].startswith("RAW_FROM_ODOO|")
    assert result["ProductMappingStatus"].tolist()[1] == "RawFromOdoo"
    assert pd.isna(result["SKU"].tolist()[1])
    assert pd.isna(result["ProductLevel"].tolist()[1])


def test_product_mapper_does_not_match_official_product_name_fallback() -> None:
    sales = pd.DataFrame(
        [{"product_name_raw": "Official Product One", "product_name_clean": "OFFICIAL PRODUCT ONE"}]
    )
    product_master = pd.DataFrame(
        [
            {
                "ProductKey": "P1",
                "ProductName": "Official Product One",
                "OdooProductName": "Different Odoo Alias",
                "SKU": "SKU1",
                "ProductLevel": "SKU",
                "Company": "A",
                "IsActive": 1,
            }
        ]
    )

    result = ProductMapper.attach(sales, product_master)

    assert result.loc[0, "ProductMappingStatus"] == "RawFromOdoo"
    assert result.loc[0, "ProductKey"].startswith("RAW_FROM_ODOO|")


def test_product_mapper_uses_same_unique_keys_as_dimension_for_duplicate_manual_keys() -> None:
    product_master = pd.DataFrame(
        [
            {
                "ProductKey": "P1",
                "ProductName": "Official One",
                "OdooProductName": "Product One",
                "Size": "",
                "SKU": "SKU1",
                "ProductLevel": "SKU",
                "Company": "A",
                "IsActive": 1,
            },
            {
                "ProductKey": "P1",
                "ProductName": "Official One Alias",
                "OdooProductName": "Product One Alias",
                "Size": "",
                "SKU": "SKU1",
                "ProductLevel": "SKU",
                "Company": "A",
                "IsActive": 1,
            },
        ]
    )
    sales = pd.DataFrame(
        [
            {"product_name_clean": "Product One", "size": ""},
            {"product_name_clean": "Product One Alias", "size": ""},
        ]
    )

    dimension = ProductDimensionBuilder().build(product_master)
    mapped = ProductMapper.attach(sales, product_master)

    assert dimension["ProductKey"].tolist() == ["P1", "P1|ROW|000002"]
    assert mapped["ProductKey"].tolist() == dimension["ProductKey"].tolist()


def test_dim_product_adds_sparse_unmapped_placeholders_for_fact_relationships() -> None:
    product_master = pd.DataFrame(
        [
            {
                "ProductKey": "P1",
                "Company": "A",
                "Category": "Cat",
                "Brand": "Brand",
                "SubBrand": "Sub",
                "Family": "Family",
                "ProductName": "Official One",
                "OdooProductName": "Odoo One",
                "ProductLevel": "SKU",
                "SKU": "SKU1",
                "Size": "S",
                "IsActive": 1,
            }
        ]
    )
    sales = ProductMapper.attach(
        pd.DataFrame(
            [
                {"product_name_raw": "Odoo One", "product_name_clean": "ODOO ONE", "company": "A"},
                {"product_name_raw": "Unknown Odoo", "product_name_clean": "UNKNOWN ODOO", "company": "A"},
            ]
        ),
        product_master,
    )

    result = ProductDimensionBuilder().build(product_master, sales)
    placeholder = result[result["ProductSource"].eq("Unmapped Odoo")].iloc[0]

    assert len(result[result["ProductSource"].eq("Input Master")]) == 1
    assert placeholder["ProductName"] == "Unknown Odoo"
    assert placeholder["OdooProductName"] == "Unknown Odoo"
    assert placeholder["Company"] == "A"
    assert bool(placeholder["IsMappedProduct"]) is False
    assert placeholder["IsActive"] == 0
    for column in ["Category", "Brand", "SubBrand", "Family", "ProductLevel", "SKU", "Size"]:
        assert pd.isna(placeholder[column])


def test_product_mapping_checks_assert_master_only_enrichment() -> None:
    product_master = pd.DataFrame(
        [{"ProductKey": "P1", "Company": "A", "ProductName": "Official One", "OdooProductName": "Odoo One"}]
    )
    sales = ProductMapper.attach(
        pd.DataFrame([{"product_name_raw": "Unknown Odoo", "product_name_clean": "UNKNOWN ODOO", "company": "A"}]),
        product_master,
    )
    dim_product = ProductDimensionBuilder().build(product_master, sales)
    qa_unmapped = QAService.unmapped_products(sales, product_master)

    checks = QAService.product_mapping_checks(dim_product, product_master, qa_unmapped).set_index("CheckName")

    assert checks.loc["Mapped products", "MetricValue"] == 1
    assert checks.loc["Unmapped Odoo products", "MetricValue"] == 1
    assert checks.loc["Dim_Product rows from input master", "MetricValue"] == 1
    assert checks.loc["Placeholder unmapped rows", "MetricValue"] == 1
    assert checks.loc["Unmapped placeholders have no enriched attributes", "Status"] == "PASS"
    assert checks.loc["Every enriched Dim_Product row exists in PRODUCTS.xlsx", "Status"] == "PASS"


def test_product_name_normalization_known_examples() -> None:
    assert normalize_product_name("متين GERMA-C2TE") == "Germa - C2TE متين"
    assert normalize_product_name("Xtreme-SuperCol S1") == "Xtreme - SuperCol S1"
    assert normalize_product_name("Xtreme XtraCol PRO") == "Xtreme XtraCol PRO"
    assert normalize_product_name(None) == "Unknown Product"


def test_product_name_columns_exist_and_are_clean_across_product_outputs() -> None:
    dim_product = ProductDimensionBuilder().build(
        pd.DataFrame(
            [
                {"ProductKey": "P1", "ProductName": "Xtreme SuperCol S1", "OdooProductName": "  Xtreme-SuperCol S1  ", "Size": ""},
                {"ProductKey": "P2", "ProductName": "Second Product", "OdooProductName": None, "Size": ""},
            ]
        )
    )
    sales = SalesCleaner().clean_product_names(
        pd.DataFrame(
            [
                {"product_name": "متين GERMA-C2TE"},
                {"product_name": "Xtreme  XtraCol PRO"},
            ]
        )
    )
    fact_sales_lines = SalesLinesFactBuilder().build(
        pd.DataFrame(
            [
                {
                    "order_number": "SO1",
                    "product_name_raw": sales.loc[0, "product_name_raw"],
                    "product_name": sales.loc[0, "product_name"],
                    "product_name_clean": sales.loc[0, "product_name_clean"],
                    "ProductNameRaw": sales.loc[0, "ProductNameRaw"],
                    "ProductName": sales.loc[0, "ProductName"],
                    "ProductNameClean": sales.loc[0, "ProductNameClean"],
                }
            ]
        )
    )
    dim_cost = ProductCostDimensionBuilder().build(
        pd.DataFrame([{"id": 1, "display_name": "Germa-C2TE", "standard_price": 1}])
    )
    fact_bcg = BCGMatrixBuilder.build(
        pd.DataFrame(
            [
                {
                    "company_final": "Majaal",
                    "SegmentKey": 10,
                    "SalesSegment": "B2B",
                    "ProductKey": "P1",
                    "product_name_clean": "متين GERMA-C2TE",
                    "order_date_date": "2026-01-02",
                    "Volume": 10,
                    "Value": 100,
                    "invoice_status": "invoiced",
                }
            ]
        ),
        as_of_date="2026-06-10",
        dim_product=dim_product,
    )

    for frame in [dim_product, fact_sales_lines, fact_bcg, dim_cost]:
        assert "ProductNameClean" in frame.columns
    assert dim_product["ProductName"].astype("string").str.strip().ne("").all()
    assert fact_bcg["ProductName"].astype("string").str.strip().ne("").all()
    assert not dim_product["ProductName"].astype("string").str.contains(r"\s{2,}", regex=True).any()
    assert not fact_bcg["ProductName"].astype("string").str.contains(r"--| -\s+- ", regex=True).any()


def test_bcg_matrix_aggregates_ytd_and_lytd_at_product_level() -> None:
    sales = pd.DataFrame(
        [
            {"company_final": "Majaal", "ProductKey": "P1", "product_name_clean": "One", "order_date_date": "2026-01-02", "Volume": 2000, "Value": 4000, "product_cost": 1},
            {"company_final": "Majaal", "ProductKey": "P1", "product_name_clean": "One", "order_date_date": "2026-02-02", "Volume": 2000, "Value": 4000, "product_cost": 1},
            {"company_final": "Majaal", "ProductKey": "P1", "product_name_clean": "One", "order_date_date": "2025-06-10", "Volume": 100, "Value": 200, "product_cost": 1},
            {"company_final": "Majaal", "ProductKey": "P1", "product_name_clean": "One", "order_date_date": "2025-06-11", "Volume": 9999, "Value": 9999, "product_cost": 1},
        ]
    )

    result = BCGMatrixBuilder.build(sales, as_of_date="2026-06-10")

    assert len(result) == 1
    assert "Period" not in result.columns
    row = result.iloc[0]
    assert row["total_quantity_YTD"] == 4000
    assert row["avg_unit_price_YTD"] == 2
    assert row["bcg_class_YTD"] == "Stars"
    assert row["total_quantity_LYTD"] == 100
    assert row["bcg_class_LYTD"] == "Strategic"
    assert row["quantity_growth_pct"] == 39
    assert row["gross_profit_change_pp"] == 0
    assert row["bcg_movement"] == "Improved"


def test_bcg_matrix_excludes_segment_columns_and_groups_by_company_product() -> None:
    sales = pd.DataFrame(
        [
            {"company_final": "Majaal", "SegmentKey": 10, "SalesSegment": "B2B", "DistributionChannel": "Retail", "ChannelKey": 1, "ProductKey": "P1", "product_name_clean": "One", "order_date_date": "2026-01-02", "Volume": 4000, "Value": 8000, "invoice_status": "invoiced"},
            {"company_final": "Majaal", "SegmentKey": 20, "SalesSegment": "B2C", "DistributionChannel": "Online", "ChannelKey": 2, "ProductKey": "P1", "product_name_clean": "One", "order_date_date": "2026-01-02", "Volume": 100, "Value": 200, "invoice_status": "to invoice"},
            {"company_final": "Majaal", "SegmentKey": 10, "SalesSegment": "B2B", "DistributionChannel": "Retail", "ChannelKey": 1, "ProductKey": "SKIP", "product_name_clean": "Skip", "order_date_date": "2026-01-02", "Volume": 100, "Value": 200, "invoice_status": "cancelled"},
        ]
    )

    result = BCGMatrixBuilder.build(sales, as_of_date="2026-06-10")

    assert not {"SegmentKey", "SalesSegment", "DistributionChannel", "ChannelKey"}.intersection(result.columns)
    assert result[["Company", "ProductKey"]].to_dict("records") == [
        {"Company": "Majaal", "ProductKey": "P1"},
    ]
    assert result["total_quantity_YTD"].sum() == 4100
    assert result["total_value_YTD"].sum() == 8200


def test_bcg_matrix_unique_by_company_product_even_with_name_variants_and_cost_join() -> None:
    sales = pd.DataFrame(
        [
            {"company_final": "Majaal", "SegmentKey": 10, "SalesSegment": "B2B", "ProductKey": "P1", "SKU": "SKU1", "product_name_clean": "One", "order_date_date": "2026-01-02", "Volume": 100, "Value": 200, "invoice_status": "invoiced"},
            {"company_final": "Majaal", "SegmentKey": 20, "SalesSegment": "B2C", "ProductKey": "P1", "SKU": "SKU1", "product_name_clean": "One Alias", "order_date_date": "2026-02-02", "Volume": 200, "Value": 400, "invoice_status": "upselling"},
        ]
    )
    costs = pd.DataFrame(
        [{"Company": "Majaal", "ProductName": "One", "SKU": "SKU1", "ProductCost": 1}]
    )

    result = BCGMatrixBuilder.build(sales, costs, as_of_date="2026-06-10")

    assert not result.duplicated(["Company", "ProductKey"]).any()
    assert len(result) == 1
    assert result["total_quantity_YTD"].sum() == 300


def test_bcg_matrix_validation_does_not_create_dim_segment_relationship() -> None:
    fact_bcg = pd.DataFrame(
        [{"Company": "Majaal", "ProductKey": "P1"}]
    )
    issues = ModelValidator.validate(
        {
            "Fact_BCGMatrix": fact_bcg,
            "Dim_Product": pd.DataFrame([{"ProductKey": "P1"}]),
            "Dim_Company": pd.DataFrame([{"CompanyKey": 1, "Company": "Majaal"}]),
            "Dim_Segment": pd.DataFrame([{"SegmentKey": 99, "Segment": "B2B"}]),
        }
    )

    assert [
        issue
        for issue in issues
        if issue.table == "Fact_BCGMatrix" and issue.check == "referential_integrity"
    ] == []


def test_bcg_matrix_product_grain_preserves_ytd_and_lytd_totals() -> None:
    sales = pd.DataFrame(
        [
            {"company_final": "Majaal", "SegmentKey": 10, "SalesSegment": "B2B", "ProductKey": "P1", "product_name_clean": "One", "order_date_date": "2026-01-02", "Volume": 10, "Value": 100, "invoice_status": "invoiced"},
            {"company_final": "Majaal", "SegmentKey": 20, "SalesSegment": "B2C", "ProductKey": "P2", "product_name_clean": "Two", "order_date_date": "2026-03-02", "Volume": 20, "Value": 200, "invoice_status": "to invoice"},
            {"company_final": "Majaal", "SegmentKey": 10, "SalesSegment": "B2B", "ProductKey": "P1", "product_name_clean": "One", "order_date_date": "2025-01-02", "Volume": 30, "Value": 300, "invoice_status": "upselling"},
            {"company_final": "Majaal", "SegmentKey": 20, "SalesSegment": "B2C", "ProductKey": "P2", "product_name_clean": "Two", "order_date_date": "2025-02-02", "Volume": 40, "Value": 400, "invoice_status": "no"},
        ]
    )

    result = BCGMatrixBuilder.build(sales, as_of_date="2026-06-10")
    summary = BCGMatrixBuilder.quality_summary(
        sales,
        result,
        as_of_date="2026-06-10",
    )

    assert result["total_value_YTD"].sum() == 300
    assert result["total_value_LYTD"].sum() == 300
    assert not result.duplicated(["Company", "ProductKey"]).any()
    assert summary["bcg_ytd_value"] == summary["fact_sales_valid_invoice_ytd_value"]
    assert summary["duplicate_company_product_count"] == 0
    assert summary["segment_columns_removed"] is True
    assert summary["ytd_lytd_columns_exist"] is True


def test_segment_dimension_contains_only_business_segments_and_normalizes_blanks() -> None:
    dim = SegmentDimensionBuilder(PipelineSettings()).build(pd.DataFrame(), pd.DataFrame())

    assert dim["Segment"].tolist() == ["B2B", "B2C", "Backoffice", "Inter Company", "Unknown"]
    assert 0 not in dim["SegmentKey"].tolist()
    assert "All" not in dim["Segment"].tolist()
    assert SegmentDimensionBuilder.normalize_segment(None) == "Unknown"
    assert SegmentDimensionBuilder.normalize_segment(" ") == "Unknown"
    assert SegmentDimensionBuilder.normalize_segment("back office") == "Backoffice"


def test_bcg_matrix_marks_new_lost_and_null_zero_lytd_growth() -> None:
    sales = pd.DataFrame(
        [
            {"company_final": "Majaal", "ProductKey": "NEW", "product_name_clean": "New", "order_date_date": "2026-01-02", "Volume": 10, "Value": 20},
            {"company_final": "Majaal", "ProductKey": "LOST", "product_name_clean": "Lost", "order_date_date": "2025-01-02", "Volume": 10, "Value": 20},
            {"company_final": "Majaal", "ProductKey": "ZERO", "product_name_clean": "Zero", "order_date_date": "2026-01-02", "Volume": 10, "Value": 20},
            {"company_final": "Majaal", "ProductKey": "ZERO", "product_name_clean": "Zero", "order_date_date": "2025-01-02", "Volume": 0, "Value": 0},
        ]
    )

    result = BCGMatrixBuilder.build(sales, as_of_date="2026-06-10").set_index("ProductKey")

    assert result.loc["NEW", "bcg_movement"] == "New"
    assert result.loc["LOST", "bcg_movement"] == "Lost"
    assert pd.isna(result.loc["NEW", "quantity_growth_pct"])
    assert pd.isna(result.loc["ZERO", "quantity_growth_pct"])


def test_product_cost_matching_priority_and_missing_cost_classification() -> None:
    costs = pd.DataFrame(
        [
            {"OdooProductID": 10, "Company": "Majaal", "ProductName": "ID Product", "SKU": "ID-SKU", "ProductCost": 6},
            {"OdooProductID": 20, "Company": "Majaal", "ProductName": "SKU Product", "SKU": "SKU-20", "ProductCost": 4},
            {"OdooProductID": 30, "Company": "Majaal", "ProductName": "Name Product", "SKU": "", "ProductCost": 3},
        ]
    )
    sales = pd.DataFrame(
        [
            {"company_final": "Majaal", "OdooProductID": 10, "ProductKey": "X", "SKU": "WRONG", "product_name_clean": "Wrong", "order_date_date": "2026-01-01", "Volume": 10, "Value": 100},
            {"company_final": "Majaal", "ProductKey": "Y", "SKU": " sku-20 ", "product_name_clean": "Wrong", "order_date_date": "2026-01-01", "Volume": 10, "Value": 100},
            {"company_final": "Majaal", "ProductKey": "Z", "SKU": "", "product_name_clean": " name   product ", "order_date_date": "2026-01-01", "Volume": 10, "Value": 100},
            {"company_final": "Majaal", "ProductKey": "MISSING", "SKU": "", "product_name_clean": "Missing", "order_date_date": "2026-01-01", "Volume": 10, "Value": 100},
        ]
    )

    attached = ProductCostMatcher.attach(sales, costs)
    assert attached["product_cost"].tolist()[:3] == [6, 4, 3]
    assert attached["ProductCostMatchKey"].str.split(":").str[0].tolist()[:3] == [
        "ID", "SKU_COMPANY", "NAME_COMPANY",
    ]

    bcg = BCGMatrixBuilder.build(sales, costs, as_of_date="2026-06-10").set_index("ProductKey")
    assert bcg.loc["X", "perc_gross_profit_YTD"] == 0.4
    assert bcg.loc["MISSING", "profit_class_YTD"] == "Unknown"
    assert bcg.loc["MISSING", "bcg_class_YTD"] == "Unclassified"
    qa = ProductCostMatcher.missing_cost_qa(sales, costs, as_of_date="2026-06-10")
    assert qa["ProductKey"].tolist() == ["MISSING"]


def test_product_cost_dimension_and_exact_bcg_columns() -> None:
    raw = pd.DataFrame(
        [{
            "id": 10,
            "product_tmpl_id_id": 100,
            "company_id": "Majaal",
            "display_name": "[SKU] Product",
            "default_code": "SKU",
            "standard_price": 5.5,
            "active": True,
        }]
    )
    dim = ProductCostDimensionBuilder().build(raw)
    assert dim.to_dict("records") == [{
        "OdooProductID": 10,
        "OdooProductTemplateID": 100,
        "Company": "Majaal",
        "ProductKey": None,
        "ProductNameRaw": "[SKU] Product",
        "ProductName": "[SKU] Product",
        "ProductNameClean": "SKU PRODUCT",
        "OdooProductName": "[SKU] Product",
        "OdooProductNameClean": "SKU PRODUCT",
        "SKU": "SKU",
        "ProductCost": 5.5,
        "IsActive": True,
    }]
    assert BCGMatrixBuilder.build(pd.DataFrame()).columns.tolist() == BCGMatrixBuilder.COLUMNS


def test_product_active_flag_reconciler_flags_products_missing_from_odoo_catalog() -> None:
    product_master = pd.DataFrame(
        [
            {"ProductKey": "P1", "IsActive": 1},
            {"ProductKey": "P1|ROW|000002", "IsActive": 1},  # alias row for the same product
            {"ProductKey": "P2", "IsActive": 1},  # archived in Odoo
            {"ProductKey": "P3", "IsActive": 1},  # no longer resolves to any Odoo record
        ]
    )
    dim_product_cost = pd.DataFrame(
        [
            {"ProductKey": "P1", "IsActive": True},
            {"ProductKey": "P2", "IsActive": False},
        ]
    )

    active_keys = ProductActiveFlagReconciler.active_base_keys(dim_product_cost)
    result = ProductActiveFlagReconciler.reconcile(product_master, active_keys, Logger(enabled=False))

    assert result.set_index("ProductKey")["IsActive"].to_dict() == {
        "P1": 1,
        "P1|ROW|000002": 1,
        "P2": 0,
        "P3": 0,
    }


def test_product_active_flag_reconciler_patches_only_input_master_rows_in_dim_product() -> None:
    dim_product = pd.DataFrame(
        [
            {"ProductKey": "P1", "IsActive": 1, "ProductSource": ProductDimensionBuilder.PRODUCT_SOURCE_INPUT},
            {"ProductKey": "P4", "IsActive": 1, "ProductSource": ProductDimensionBuilder.PRODUCT_SOURCE_INPUT},
            {"ProductKey": "P5", "IsActive": 1, "ProductSource": ProductDimensionBuilder.PRODUCT_SOURCE_ODOO},
        ]
    )

    patched = ProductActiveFlagReconciler.patch_dim_product(
        dim_product, {"P1"}, ProductDimensionBuilder.PRODUCT_SOURCE_INPUT
    )

    assert patched.set_index("ProductKey")["IsActive"].to_dict() == {"P1": 1, "P4": 0, "P5": 1}


def test_product_active_flag_reconciler_writes_back_to_products_xlsx_without_touching_other_columns(
    tmp_path: Path,
) -> None:
    import openpyxl

    path = tmp_path / "PRODUCTS.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ProductKey", "SKU", "IsActive"])
    ws.append(["P1", "SKU-1", 1])
    ws.append(["P2", "SKU-2", 1])
    ws["B3"] = "=UPPER(A3)"  # unrelated formula must survive the write-back
    wb.save(path)

    product_master = pd.DataFrame(
        [
            {"ProductKey": "P1", "IsActive": 1},
            {"ProductKey": "P2", "IsActive": 1},
        ]
    )
    active_keys = {"P1"}
    ProductActiveFlagReconciler.reconcile(product_master, active_keys, Logger(enabled=False), products_path=path)

    saved = openpyxl.load_workbook(path)
    saved_ws = saved.active
    assert [cell.value for cell in saved_ws[2]] == ["P1", "SKU-1", 1]
    assert saved_ws["A3"].value == "P2"
    assert saved_ws["C3"].value == 0
    assert saved_ws["B3"].value == "=UPPER(A3)"


def test_workbook_exporter_writes_product_qa_separately_and_removes_old_ambiguous_file(tmp_path: Path) -> None:
    ambiguous = tmp_path / "QA_AmbiguousProductMapping.xlsx"
    ambiguous.write_text("old", encoding="utf-8")
    sheets = {
        "Dim_Product": pd.DataFrame([{"ProductKey": "P1"}]),
        "QA_UnmappedProducts": pd.DataFrame([{"product_name_clean": "Missing", "size": ""}]),
    }

    WorkbookExporter(type("Logger", (), {"ok": lambda self, message: None})()).export(
        sheets,
        tmp_path / "SalesModel_OneOutput.xlsx",
    )

    assert not ambiguous.exists()
    assert (tmp_path / "QA_UnmappedProducts.xlsx").exists()
    assert pd.ExcelFile(tmp_path / "SalesModel_OneOutput.xlsx").sheet_names == ["Dim_Product"]


def test_workbook_exporter_writes_alternate_when_locked_workbook_save_fails(tmp_path: Path, monkeypatch) -> None:
    model_path = tmp_path / "SalesModel_OneOutput.xlsx"
    alternate_path = tmp_path / "SalesModel_OneOutput_unlocked.xlsx"
    sheets = {
        "Dim_Product": pd.DataFrame([{"ProductKey": "P1"}]),
        "QA_UnmappedProducts": pd.DataFrame([{"OdooProductName": "Missing"}]),
    }
    original_write = WorkbookExporter._write_workbook_file

    def locked_primary_write(self, path, sheets, existing_sheet_order, preserved_sheets, separate_qa_sheets):
        if Path(path) == model_path:
            raise PermissionError("locked")
        return original_write(self, path, sheets, existing_sheet_order, preserved_sheets, separate_qa_sheets)

    monkeypatch.setattr(WorkbookExporter, "_write_workbook_file", locked_primary_write)
    monkeypatch.setattr(WorkbookExporter, "_export_to_open_excel", lambda self, sheets, path: (_ for _ in ()).throw(RuntimeError("sharing violation")))
    monkeypatch.setattr(WorkbookExporter, "_alternate_workbook_path", staticmethod(lambda path: alternate_path))

    logger = type("Logger", (), {"ok": lambda self, message: None, "warn": lambda self, message: None})()
    result = WorkbookExporter(logger).export(sheets, model_path)

    assert result == alternate_path
    assert alternate_path.exists()
    assert pd.ExcelFile(alternate_path).sheet_names == ["Dim_Product"]
    assert (tmp_path / "QA_UnmappedProducts.xlsx").exists()


def test_bcg_refresh_upsert_preserves_existing_workbook_sheets(tmp_path: Path) -> None:
    from scripts.refresh_product_outputs import _upsert_model_sheets

    model_path = tmp_path / "SalesModel_OneOutput.xlsx"
    with pd.ExcelWriter(model_path) as writer:
        pd.DataFrame([{"Keep": "yes"}]).to_excel(writer, sheet_name="ExistingSheet", index=False)
        pd.DataFrame([{"Old": 1}]).to_excel(writer, sheet_name="Fact_BCGMatrix", index=False)
        pd.DataFrame([{"SegmentKey": 1, "Segment": "B2B"}]).to_excel(writer, sheet_name="Dim_Segment", index=False)

    _upsert_model_sheets(
        model_path,
        {
            "Fact_BCGMatrix": pd.DataFrame([{"Company": "Majaal", "ProductKey": "P1"}]),
        },
    )

    book = pd.ExcelFile(model_path)
    assert book.sheet_names == ["ExistingSheet", "Fact_BCGMatrix", "Dim_Segment"]
    assert pd.read_excel(book, sheet_name="ExistingSheet").to_dict("records") == [{"Keep": "yes"}]
    assert pd.read_excel(book, sheet_name="Fact_BCGMatrix").to_dict("records") == [
        {"Company": "Majaal", "ProductKey": "P1"}
    ]
    assert pd.read_excel(book, sheet_name="Dim_Segment").to_dict("records") == [
        {"SegmentKey": 1, "Segment": "B2B"},
    ]


def test_workbook_exporter_datetime_serial_preserves_local_clock_time() -> None:
    assert WorkbookExporter._excel_cell_value(pd.Timestamp("2026-06-16 13:09:37")) == (
        pd.Timestamp("2026-06-16 13:09:37") - WorkbookExporter.EXCEL_EPOCH
    ) / pd.Timedelta(days=1)


def test_workbook_exporter_datetime_detector_handles_string_dtype() -> None:
    assert WorkbookExporter._is_datetime_series(pd.Series(["B2B", "B2C"], dtype="string")) is False


def test_excel_sql_validation_reads_unmapped_products_from_separate_workbook(
    tmp_path: Path,
    monkeypatch,
) -> None:
    model_path = tmp_path / "SalesModel_OneOutput.xlsx"
    with pd.ExcelWriter(model_path) as writer:
        pd.DataFrame(columns=["order_number", "OrderDateTime"]).to_excel(writer, sheet_name="Fact_Orders", index=False)
        pd.DataFrame(columns=["order_number", "order_date", "order_date_date"]).to_excel(writer, sheet_name="Fact_SalesLines", index=False)
    pd.DataFrame([{"product_name_clean": "Missing", "size": ""}]).to_excel(
        tmp_path / "QA_UnmappedProducts.xlsx",
        index=False,
    )

    settings = Settings(
        odoo_url="x",
        odoo_db="x",
        odoo_user="x",
        odoo_api_key="x",
        input_dir=tmp_path,
        output_dir=tmp_path,
        output_file=model_path.name,
        batch_size=100,
        timezone="Africa/Tripoli",
        assume_utc_for_naive=False,
    )
    pipeline = PowerBISalesPipeline(settings)
    exporter = SimpleNamespace(
        _count_rows=lambda table_name: 1 if table_name == "QA_UnmappedProducts" else 0,
        _effective_schema=lambda: None,
    )
    monkeypatch.setattr(pipeline, "_latest_sql_tuple_for_table", lambda *args, **kwargs: (None, None))
    monkeypatch.setattr(pipeline, "_latest_sql_sales_line_tuple", lambda *args, **kwargs: (None, None, None))

    pipeline._validate_excel_matches_sql_output(
        model_path,
        exporter,
        {
            "Fact_Orders": pd.DataFrame(),
            "Fact_SalesLines": pd.DataFrame(),
            "QA_UnmappedProducts": pd.DataFrame([{"product_name_clean": "Missing", "size": ""}]),
        },
    )
